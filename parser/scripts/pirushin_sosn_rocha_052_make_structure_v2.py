#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pirushin_sosn_rocha_052_make_structure_v1.py

ОСВ (счёт 01, 1С) → единая иерархическая JSON-структура:

    Предприятие
      └─ Бизнес-единицы (структурные подразделения)
           └─ Кадастровые объекты (ЗУ / Здание / Сооружение / Помещение / ОНС)
                └─ Уровни (этажи, подвал, мансарда, антресоль …)
                     └─ Оборудование (основные средства из ОСВ)

Источники:
  • ОСВ (.xlsx) по счёту 01 из 1С                    — обязательный вход.
  • Предыдущий structure_*.json                       — опц. база для merge.
  • Каталог JSON-выгрузок NSPD/ЕГРН (от 01_parsing*)   — опц. адреса/этажность.

Гарантии:
  • Идемпотентность: детерминированные ID (sha1(account|name|…)). Повторный
    запуск с тем же входом → тот же набор ID. Поля с confirmed=true никогда
    не перезаписываются.
  • Частичные данные: все обогащаемые поля optional, есть статусы (stub /
    partial / complete).
  • Безопасность: парсинг устойчив к 01.К (Cyrillic К), к пустым строкам, к
    отсутствию NSPD-данных. Никаких сетевых запросов.

Требования: Python 3.13+, Windows 10, openpyxl.
Запуск: python pirushin_sosn_rocha_052_make_structure_v1.py
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ════════════════════════════════════════════════════════════════════════════
#  Цветной вывод (Windows-совместимо)
# ════════════════════════════════════════════════════════════════════════════
class C:
    G  = "\033[92m"
    R  = "\033[91m"
    Y  = "\033[93m"
    CY = "\033[96m"
    M  = "\033[95m"
    B  = "\033[1m"
    X  = "\033[0m"


def cp(text: str = "", color: str = "") -> None:
    print(f"{color}{text}{C.X}" if color else text)


# ════════════════════════════════════════════════════════════════════════════
#  Регулярные выражения и колонки ОСВ
# ════════════════════════════════════════════════════════════════════════════
ACCOUNT_RE      = re.compile(r"^\d{2}\.(?:\d{2,3}|К)$")
CN_FULL_RE      = re.compile(r"\b\d{2}:\d{2}:\d{2,8}:\d{1,8}(?:/\d+)?\b")
CN_TAIL_RE      = re.compile(r"(?<!\d):(\d{2,8})\b")
INV_RE          = re.compile(r"\b01[_\.]\d{2,5}(?:[_\.]\d{2,5})?\b")
DATE_RE         = re.compile(r"\b(\d{2}\.\d{2}\.\d{4})\b")
CONTRACT_NO_RE  = re.compile(r"№\s*([\w\-/]+)", re.IGNORECASE)
QUARTER_RE      = re.compile(r"(\d)\s*квартал[а-я]*\s*(\d{4})", re.IGNORECASE)
YEAR_RE         = re.compile(r"за\s*(\d{4})\s*г", re.IGNORECASE)
DATE_RANGE_RE   = re.compile(r"за\s*(\d{2}\.\d{2}\.\d{4})\s*[-–—]\s*(\d{2}\.\d{2}\.\d{4})")

# Имена меток в col 1 (БУ/НУ/ВР) и служебные строки
SECTION_LABELS  = {"БУ", "НУ", "ВР", "Арендованные ОС", "Основные средства", "Итого"}

# Структура колонок ОСВ
COL_NAME, COL_LABEL = 0, 1
COL_OPEN_DT, COL_OPEN_KT       = 2, 3
COL_TURN_DT, COL_TURN_KT       = 4, 5
COL_CLOSE_DT, COL_CLOSE_KT     = 6, 7

AMOUNT_KEYS = ("open_dt", "open_kt", "turnover_dt", "turnover_kt", "close_dt", "close_kt")
ZERO_AMOUNTS: dict[str, float] = {k: 0.0 for k in AMOUNT_KEYS}


# ════════════════════════════════════════════════════════════════════════════
#  Справочники (нормализованные)
# ════════════════════════════════════════════════════════════════════════════
OPF_MAP: dict[str, str] = {
    "ООО":  "Общество с ограниченной ответственностью",
    "АО":   "Акционерное общество",
    "ПАО":  "Публичное акционерное общество",
    "ОАО":  "Открытое акционерное общество",
    "ЗАО":  "Закрытое акционерное общество",
    "ИП":   "Индивидуальный предприниматель",
    "ГУП":  "Государственное унитарное предприятие",
    "МУП":  "Муниципальное унитарное предприятие",
    "АНО":  "Автономная некоммерческая организация",
    "ФГБУ": "Федеральное государственное бюджетное учреждение",
    "ФГУП": "Федеральное государственное унитарное предприятие",
    "НКО":  "Некоммерческая организация",
    "ТСЖ":  "Товарищество собственников жилья",
}

ACCOUNT_MAP: dict[str, dict[str, Any]] = {
    "01":    {"label": "Основные средства (свод)",                  "right": None},
    "01.01": {"label": "Основные средства на балансе организации",  "right": "собственность"},
    "01.03": {"label": "Арендованные основные средства",            "right": "аренда"},
    "01.09": {"label": "Амортизация арендованных ОС",               "right": "аренда"},
    "01.К":  {"label": "Арендные платежи (накопление)",             "right": "аренда"},
}

# Эвристика для подсказки бизнес-единиц (используется только когда адрес не
# определён, см. Phase B в build_business_units).
BU_KEYWORDS: dict[str, list[str]] = {
    "Медицина / Лечение": [
        "аппарат", "криотерап", "ванна", "реабокс", "массаж", "ингал",
        "ультразвук", "электрофор", "магнит", "лазер", "стоматолог",
        "тренажёр медиц", "процедур",
    ],
    "Спа / Бассейн / Бани": [
        "сауна", "хаммам", "парная", "бассейн", "джакузи", "купель",
        "пресотерап", "softcooker",
    ],
    "Ресторан / Кухня / Бар": [
        "плита", "пароконвектомат", "холодильник", "морозильник", "пицц",
        "соус", "фритюр", "посудомоечн", "кофе", "напитков", "бар",
        "аквариум для устриц", "акваферма",
    ],
    "Спорт / Фитнес": [
        "эллипсоид", "беговая дорожка", "велотренажёр", "штанга",
        "гантел", "тренаж", "precor",
    ],
    "Жилые номера / Гостиница": [
        "торшер", "кровать", "шкаф", "тумба", "матрас", "комод",
        "прикроват", "телевизор", "минибар",
    ],
    "Офис / Управление": [
        "xerox", "принтер", "мфу", "сканер", "ноутбук", "монитор",
        "компьютер", "проектор",
    ],
    "Транспорт / Водная техника": [
        "автомобиль", "транспортное средство", "ford", "lada", "toyota",
        "гидроцикл", "лодка", "катер", "яхта", "квадроцикл",
    ],
    "Земля / Ландшафт": [
        "земельный участок", "ландшафт", "благоустр", "тротуар",
    ],
    "Религиозное / Декор": [
        "аналой", "икон", "свечн", "храм", "церков", "электрокамин",
    ],
    "Технические / Инженерия": [
        "генератор", "трансформатор", "котёл", "котел", "насос",
        "кондиционер", "вентил",
    ],
}

# Признаки «уличного» расположения (бассейн, генератор, лодка → land_point)
OUTDOOR_HINTS: tuple[str, ...] = (
    "генератор", "трансформатор", "бассейн", "гидроцикл", "лодка", "катер",
    "яхта", "транспортное средство", "автомобиль", "акваферма",
)

# Справочник уровней (по СП 54.13330.2016 + практике ЕГРН)
LEVEL_TYPES: list[dict[str, str]] = [
    {"key": "floor",              "name": "Этаж"},
    {"key": "mansarda",           "name": "Мансарда"},
    {"key": "mezonin",            "name": "Мезонин"},
    {"key": "basement",           "name": "Подвал"},
    {"key": "ground_floor",       "name": "Цокольный этаж"},
    {"key": "added_floor",        "name": "Надстроенный этаж"},
    {"key": "technical_floor",    "name": "Технический этаж"},
    {"key": "attic",              "name": "Чердак"},
    {"key": "antresol",           "name": "Антресоль"},
    {"key": "svetelka",           "name": "Светёлка"},
    {"key": "semi_basement",      "name": "Полуподвал"},
    {"key": "antresol_basement",  "name": "Антресоль подвала"},
    {"key": "antresol_ground",    "name": "Антресоль цокольного этажа"},
    {"key": "attic_addon",        "name": "Чердачная надстройка"},
    {"key": "operable_roof",      "name": "Эксплуатируемая кровля"},
    {"key": "tech_underfloor",    "name": "Техническое подполье"},
    {"key": "interfloor",         "name": "Междуэтажное пространство"},
]

CATEGORY_TO_OBJ_TYPE: dict[str, str] = {
    "Земельные участки":                          "Земельный участок",
    "Здания":                                     "Здание",
    "Сооружения":                                 "Сооружение",
    "Помещения":                                  "Помещение",
    "Объекты незавершенного строительства":       "Объект незавершенного строительства",
}


# ════════════════════════════════════════════════════════════════════════════
#  Утилиты
# ════════════════════════════════════════════════════════════════════════════
def sha1_short(s: str, n: int = 8) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:n]


def slugify(s: str) -> str:
    """ASCII-slug для имён файлов и ID. Кириллица транслитерируется."""
    table = {
        "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"zh",
        "з":"z","и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o",
        "п":"p","р":"r","с":"s","т":"t","у":"u","ф":"f","х":"h","ц":"c",
        "ч":"ch","ш":"sh","щ":"sch","ъ":"","ы":"y","ь":"","э":"e","ю":"yu",
        "я":"ya",
    }
    out: list[str] = []
    for ch in s.lower():
        if ch.isalnum() and ord(ch) < 128:
            out.append(ch)
        elif ch in table:
            out.append(table[ch])
        elif ch.isspace() or ch in "-_":
            out.append("_")
    res = re.sub(r"_+", "_", "".join(out)).strip("_")
    return res or "unknown"


def norm_ws(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace(" ", " ")).strip()


def to_num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def ask(prompt: str, default: str = "") -> str:
    suffix = f" [{default}]" if default else ""
    ans = input(f"{prompt}{suffix}: ").strip().strip('"').strip("'")
    return ans or default


def ask_yn(prompt: str, default: bool = True) -> bool:
    suf = " [Y/n]" if default else " [y/N]"
    a = input(f"{prompt}{suf}: ").strip().lower()
    return default if not a else a in ("y", "yes", "д", "да")


# ════════════════════════════════════════════════════════════════════════════
#  Извлечение enterprise из шапки ОСВ
# ════════════════════════════════════════════════════════════════════════════
def _parse_period(period_label: str) -> dict[str, Any]:
    """Распознаёт: 'X квартал YYYY', 'за YYYY г.', 'за DD.MM.YYYY - DD.MM.YYYY'."""
    period: dict[str, Any] = {"label": period_label, "from": None, "to": None}

    if (m := QUARTER_RE.search(period_label)):
        q, year = int(m.group(1)), int(m.group(2))
        start_m = (q - 1) * 3 + 1
        end_m = start_m + 2
        last_day = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][end_m - 1]
        if end_m == 2 and (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)):
            last_day = 29
        period["from"] = f"{year}-{start_m:02d}-01"
        period["to"]   = f"{year}-{end_m:02d}-{last_day:02d}"
        return period

    if (m := DATE_RANGE_RE.search(period_label)):
        d1 = datetime.strptime(m.group(1), "%d.%m.%Y").date()
        d2 = datetime.strptime(m.group(2), "%d.%m.%Y").date()
        period["from"] = d1.isoformat()
        period["to"]   = d2.isoformat()
        return period

    if (m := YEAR_RE.search(period_label)):
        year = int(m.group(1))
        period["from"] = f"{year}-01-01"
        period["to"]   = f"{year}-12-31"

    return period


def parse_enterprise(header_row: str, period_row: str) -> dict[str, Any]:
    h = norm_ws(header_row)
    opf, opf_full = None, None
    for key in sorted(OPF_MAP, key=len, reverse=True):
        full = OPF_MAP[key]
        if h.lower().startswith(full.lower()) or re.search(rf"\b{re.escape(key)}\b", h):
            opf, opf_full = key, full
            break

    # Имя организации: всё, что после OPF, без обрамляющих внешних кавычек
    name_short = h
    if opf_full and h.lower().startswith(opf_full.lower()):
        name_short = h[len(opf_full):].strip()
    elif opf and (m := re.search(rf"\b{re.escape(opf)}\b\s*(.+)", h)):
        name_short = m.group(1).strip()
    # Снимаем только парные внешние кавычки, внутренние сохраняем
    for pair in ('""', "««", "»»", "«»"):
        pass  # placeholder for explicit reading; logic below handles all kinds
    while len(name_short) >= 2 and name_short[0] in '"«' and name_short[-1] in '"»':
        name_short = name_short[1:-1].strip()

    slug = slugify(name_short or h)
    return {
        "id": f"ent_{slug}",
        "slug": slug,
        "opf": opf,
        "opf_full": opf_full,
        "name_short": name_short,
        "name_full": h,
        "inn": None,
        "ogrn": None,
        "kpp": None,
        "period": _parse_period(norm_ws(period_row)),
    }


# ════════════════════════════════════════════════════════════════════════════
#  Парсинг ОСВ
# ════════════════════════════════════════════════════════════════════════════
def _row_amounts(row: tuple) -> dict[str, float]:
    cells = list(row) + [None] * max(0, COL_CLOSE_KT + 1 - len(row))
    return {
        "open_dt":     to_num(cells[COL_OPEN_DT]),
        "open_kt":     to_num(cells[COL_OPEN_KT]),
        "turnover_dt": to_num(cells[COL_TURN_DT]),
        "turnover_kt": to_num(cells[COL_TURN_KT]),
        "close_dt":    to_num(cells[COL_CLOSE_DT]),
        "close_kt":    to_num(cells[COL_CLOSE_KT]),
    }


def _extract_cadastral_hints(name: str) -> list[str]:
    hints: list[str] = []
    seen: set[str] = set()
    for m in CN_FULL_RE.findall(name):
        if m not in seen:
            seen.add(m); hints.append(m)
    for m in CN_TAIL_RE.findall(name):
        token = f":{m}"
        if token not in seen:
            seen.add(token); hints.append(token)
    return hints


def parse_osv(xlsx_path: Path) -> tuple[dict, list[dict]]:
    """
    Возвращает (enterprise, equipment_records).

    equipment record:
      { account, name, inv_hint, cadastral_hints, contract_date, contract_no,
        right_type, amounts: {bu: {...}, nu: {...}, vr: {...}} }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header = ""
    period_label = ""
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            header = norm_ws(row[0])
        elif i == 2:
            period_label = norm_ws(row[0])
        rows.append(row)

    enterprise = parse_enterprise(header, period_label)

    items: list[dict] = []
    current_account: str | None = None
    current: dict | None = None

    for row in rows:
        if not row:
            continue
        c0 = norm_ws(row[COL_NAME])
        c1 = norm_ws(row[COL_LABEL]) if len(row) > COL_LABEL else ""
        if not c0 and not c1:
            continue

        if c0 and ACCOUNT_RE.match(c0):
            current_account = c0
            current = None
            continue
        if c0 == "Итого":
            current = None
            continue

        # Начало тройки строк (БУ + НУ + ВР)
        if c0 and c1 == "БУ" and c0 not in SECTION_LABELS:
            if current_account is None:
                continue
            current = {
                "account": current_account,
                "name": c0,
                "amounts": {
                    "bu": _row_amounts(row),
                    "nu": dict(ZERO_AMOUNTS),
                    "vr": dict(ZERO_AMOUNTS),
                },
            }
            items.append(current)
            continue

        if current is not None and not c0 and c1 in ("НУ", "ВР"):
            current["amounts"]["nu" if c1 == "НУ" else "vr"] = _row_amounts(row)

    # Post-processing: подсказки из названия
    for it in items:
        name = it["name"]
        it["cadastral_hints"] = _extract_cadastral_hints(name)
        it["inv_hint"]        = (m.group(0) if (m := INV_RE.search(name)) else None)
        it["contract_date"]   = (m.group(1) if (m := DATE_RE.search(name)) else None)
        it["contract_no"]     = (m.group(1) if (m := CONTRACT_NO_RE.search(name)) else None)
        it["right_type"]      = (ACCOUNT_MAP.get(it["account"], {}) or {}).get("right")

    return enterprise, items


# ════════════════════════════════════════════════════════════════════════════
#  Загрузка JSON-источников
# ════════════════════════════════════════════════════════════════════════════
def load_json_safe(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        cp(f"  ⚠ Не удалось прочитать {path.name}: {e}", C.Y)
        return None


def load_nspd_objects(folder: Path) -> list[dict]:
    """
    Читает каталог *.json (выгрузки 01_parsing_nspd_gemini_v5.py / enrich_v11
    / egrn_parser). Поддерживает три формата:
      • nspd-style    : {"data": {"Категория": {"КН": {…info…}}}}
      • enriched-style: то же + ключи beneficiaries / business_units (последние
                         обрабатываются load_enriched_extras).
      • egrn-style    : {"tables": {"building_objects": [{cad_number, …}],
                                     "land_objects":     [{…}]}}
    Возвращает плоский список: [{cn, category, info, related, sources_meta}, …].
    """
    if not folder.exists() or not folder.is_dir():
        return []
    out: list[dict] = []
    for jp in folder.rglob("*.json"):
        data = load_json_safe(jp)
        if not isinstance(data, dict):
            continue
        out.extend(_extract_from_nspd_style(data, jp))
        out.extend(_extract_from_egrn_style(data, jp))
    return out


def _extract_from_nspd_style(data: dict, source: Path) -> list[dict]:
    payload = data["data"] if isinstance(data.get("data"), dict) else data
    out: list[dict] = []
    for category, by_cn in (payload or {}).items():
        if not isinstance(by_cn, dict) or category in ("beneficiaries", "business_units", "founder_chains"):
            continue
        for cn, info in by_cn.items():
            if not isinstance(info, dict):
                continue
            related = info.get("Связанные объекты", {})
            out.append({
                "cn": cn,
                "category": category,
                "info": {k: v for k, v in info.items() if k != "Связанные объекты"},
                "related": related if isinstance(related, dict) else {},
                "sources_meta": {"file": source.name, "kind": "nspd|enriched"},
            })
    return out


def _extract_from_egrn_style(data: dict, source: Path) -> list[dict]:
    tables = data.get("tables") if isinstance(data.get("tables"), dict) else None
    if not tables:
        return []
    out: list[dict] = []
    for tbl_name, category in (("building_objects", "Здания"),
                                ("land_objects",     "Земельные участки")):
        for row in tables.get(tbl_name, []) or []:
            if not isinstance(row, dict) or not row.get("cad_number"):
                continue
            obj_type = row.get("object_type") or category
            mapped_category = {
                "land":     "Земельные участки",
                "building": "Здания",
                "room":     "Помещения",
                "ons":      "Объекты незавершенного строительства",
            }.get(str(obj_type).lower(), category)
            info = {
                "Вид объекта недвижимости": obj_type,
                "Адрес":            row.get("address"),
                "Площадь, кв. м":   row.get("area"),
                "Кадастровая стоимость, руб.": row.get("cadastral_value"),
                "Количество этажей": row.get("floors_above_ground") or row.get("floors_total"),
                "Количество подземных этажей": row.get("underground_floors"),
                "Назначение":       row.get("purpose"),
                "Наименование":     row.get("name"),
                "Дата постановки на учёт": row.get("registration_date"),
                "Кадастровый квартал":     row.get("quarter_cad_number"),
            }
            info = {k: v for k, v in info.items() if v not in (None, "")}
            out.append({
                "cn": row["cad_number"],
                "category": mapped_category,
                "info": info,
                "related": {},
                "sources_meta": {"file": source.name, "kind": "egrn"},
            })
    return out


def load_enriched_extras(folder: Path) -> dict[str, Any]:
    """
    Дополнительно извлекает из enriched_*.json:
      • beneficiaries (для ИНН/ОГРН/КПП по бизнес-единицам)
      • business_units (для cross-link external_id)
    Возвращает {"beneficiaries": {...}, "business_units": [...]}.
    """
    out: dict[str, Any] = {"beneficiaries": {}, "business_units": []}
    if not folder.exists() or not folder.is_dir():
        return out
    for jp in folder.rglob("*.json"):
        data = load_json_safe(jp)
        if not isinstance(data, dict):
            continue
        payload = data.get("data") if isinstance(data.get("data"), dict) else data
        if isinstance(payload, dict):
            ben = payload.get("beneficiaries") or data.get("beneficiaries")
            if isinstance(ben, dict):
                for k, v in ben.items():
                    if isinstance(v, dict):
                        out["beneficiaries"][k] = v
            bus = data.get("business_units") or payload.get("business_units")
            if isinstance(bus, list):
                for b in bus:
                    if isinstance(b, dict):
                        out["business_units"].append(b)
    return out


# ════════════════════════════════════════════════════════════════════════════
#  Кадастровые объекты
# ════════════════════════════════════════════════════════════════════════════
def build_cadastre_objects(
    equipment_records: list[dict],
    nspd_items: list[dict],
) -> tuple[list[dict], dict[str, str]]:
    """
    Сначала из NSPD/ЕГРН (полные КН + атрибуты), затем добираем заглушки из
    подсказок ОСВ. Возвращает (cadastre_objects, hint→cad_id).
    """
    cad_by_cn: dict[str, dict] = {}
    for ni in nspd_items:
        cn = ni["cn"]
        if cn in cad_by_cn:
            continue
        obj_type = CATEGORY_TO_OBJ_TYPE.get(ni["category"], "Объект")
        info = ni["info"]
        cad_by_cn[cn] = {
            "id": f"cad_{sha1_short(cn)}",
            "cadastral_number": cn,
            "object_type": info.get("Вид объекта недвижимости") or obj_type,
            "right_type": None,
            "address": info.get("Адрес") or info.get("Местоположение"),
            "area": (
                info.get("Площадь, кв.м")
                or info.get("Площадь, кв. м")
                or info.get("Площадь общая")
                or info.get("Площадь уточненная")
                or info.get("Площадь")
            ),
            "parent_cadastral": None,
            "levels": [],
            "enrichment_status": "partial",
            "sources": ["nspd"],
            "_raw_info": info,
            "_geometry": info.get("_geometry") if isinstance(info.get("_geometry"), dict) else None,
        }

    hint_to_id: dict[str, str] = {}
    for rec in equipment_records:
        for hint in rec["cadastral_hints"]:
            key: str = hint
            if not CN_FULL_RE.fullmatch(hint):
                tail = hint.lstrip(":")
                for known in cad_by_cn:
                    if known.endswith(":" + tail):
                        key = known
                        break
            if key in cad_by_cn:
                hint_to_id[hint] = cad_by_cn[key]["id"]
                continue
            cad_by_cn[key] = {
                "id": f"cad_{sha1_short(key)}",
                "cadastral_number": key,
                "object_type": "Земельный участок" if "земельн" in rec["name"].lower() else "Объект",
                "right_type": (ACCOUNT_MAP.get(rec["account"], {}) or {}).get("right"),
                "address": None,
                "area": None,
                "parent_cadastral": None,
                "levels": [],
                "enrichment_status": "stub",
                "sources": ["osv_hint"],
            }
            hint_to_id[hint] = cad_by_cn[key]["id"]
    return list(cad_by_cn.values()), hint_to_id


def compute_levels_for_cadastre(cad: dict) -> list[dict]:
    """Авто-генерация уровней (снизу вверх) из ЕГРН-атрибутов 'Количество этажей'."""
    info = cad.get("_raw_info") or {}
    if "земельный" in (cad.get("object_type") or "").lower():
        return []

    def to_int(v: Any) -> int:
        if not v:
            return 0
        m = re.search(r"\d+", str(v))
        return int(m.group(0)) if m else 0

    above_n = to_int(
        info.get("Количество этажей (в том числе подземных)")
        or info.get("Количество этажей")
        or info.get("Этажность")
        or info.get("floors_above_ground")
    )
    below_n = to_int(info.get("Количество подземных этажей") or info.get("underground_floors"))

    levels: list[dict] = []
    n = 1
    cn = cad["cadastral_number"]
    for i in range(below_n, 0, -1):
        kind = "Подвал" if i > 1 else "Цокольный этаж"
        levels.append({
            "id": f"lvl_{sha1_short(f'{cn}_b{i}')}",
            "number": n,
            "type": kind,
            "label": f"Уровень {n}. {kind} {i}",
            "underground": True,
            "cadastral_source": cn,
            "confirmed": False,
        })
        n += 1
    for i in range(1, above_n + 1):
        levels.append({
            "id": f"lvl_{sha1_short(f'{cn}_a{i}')}",
            "number": n,
            "type": "Этаж",
            "label": f"Уровень {n}. Этаж {i}",
            "underground": False,
            "cadastral_source": cn,
            "confirmed": False,
        })
        n += 1
    return levels


# ════════════════════════════════════════════════════════════════════════════
#  Бизнес-единицы (двухфазная стратегия)
# ════════════════════════════════════════════════════════════════════════════
def _address_short(addr: str | None) -> str | None:
    if not addr:
        return None
    a = norm_ws(addr)
    a = re.sub(r"\b\d{6}\b,?\s*", "", a)
    a = re.sub(r"^Россий[а-я]+ Федерация,?\s*", "", a, flags=re.I)
    a = re.sub(r"^Республика [^,]+,?\s*", "", a, flags=re.I)
    a = re.sub(r"\bобл[а-я]*\.?\s+[^,]+,\s*", "", a, flags=re.I)
    return a[:80] or None


def _classify_bu_heuristic(name: str) -> str:
    nl = name.lower()
    for bu, kws in BU_KEYWORDS.items():
        if any(kw in nl for kw in kws):
            return bu
    return "Прочее / Не классифицировано"


def _is_outdoor(name: str) -> bool:
    nl = name.lower()
    return any(h in nl for h in OUTDOOR_HINTS)


def _resolve_equipment_address(
    rec: dict,
    cadastre_by_id: dict[str, dict],
    hint_to_cad_id: dict[str, str],
) -> tuple[str | None, str | None]:
    """Возвращает (cad_id, address) для оборудования или (None, None)."""
    for hint in rec["cadastral_hints"]:
        cid = hint_to_cad_id.get(hint)
        if not cid:
            continue
        cad = cadastre_by_id.get(cid)
        if cad and cad.get("address"):
            return cid, cad["address"]
        if cad:
            return cid, None
    return None, None


def build_business_units(
    records: list[dict],
    cadastre_objects: list[dict],
    hint_to_cad_id: dict[str, str],
) -> tuple[list[dict], dict[str, str]]:
    """
    Фаза А: группировка по адресам кадастровых объектов (из NSPD/ЕГРН).
    Фаза Б: оборудование без адреса — интерактивно (имя/skip/attach to bu_id),
            эвристика по ключевым словам используется только как подсказка.

    Возвращает (business_units, eq_id → bu_id).
    """
    cadastre_by_id = {c["id"]: c for c in cadastre_objects}
    addr_groups: dict[str, list[dict]] = {}
    eq_unassigned: list[dict] = []

    for rec in records:
        if rec["account"] == "01.К":  # арендные платежи — не в BU, попадут в lease_contracts
            continue
        _, addr = _resolve_equipment_address(rec, cadastre_by_id, hint_to_cad_id)
        (addr_groups.setdefault(addr, []) if addr else eq_unassigned).append(rec)

    bus: list[dict] = []
    eq_to_bu: dict[str, str] = {}

    # ─── Phase A: address-based ─────────────────────────────────────────
    if addr_groups:
        cp("\n" + "═" * 70, C.B)
        cp("Фаза А. Бизнес-единицы по адресам кадастровых объектов", C.B)
        cp("═" * 70, C.B)
        cp("Для каждой группы — Enter оставить предложенное имя или ввести своё.\n")
        for addr, items in sorted(addr_groups.items(), key=lambda kv: -len(kv[1])):
            short = _address_short(addr) or addr
            cp(f"  Адрес: {short}", C.CY)
            cp(f"  Объектов ОС в этой группе: {len(items)}")
            cp("  Примеры: " + " · ".join(it["name"][:40] for it in items[:4]))
            name = ask("  Имя бизнес-единицы", default=short)
            # Якорь — самый «младший» (числовой) полный КН группы, чтобы ID BU
            # был стабильным между запусками даже при правке адреса/имени.
            anchor_cn = sorted({
                h for it in items for h in it["cadastral_hints"]
                if CN_FULL_RE.fullmatch(h)
            }) or [addr]
            bu_id = f"bu_{sha1_short(anchor_cn[0])}"
            bus.append({
                "id": bu_id,
                "name": name,
                "keywords": [],
                "address": addr,
                "anchor_cadastral": anchor_cn[0] if anchor_cn else None,
                "cadastrals": anchor_cn if isinstance(anchor_cn, list) and anchor_cn != [addr] else [],
                "parent_id": None,
                "external_ids": {},
                "source": "address",
                "confirmed": True,
            })
            for it in items:
                eq_to_bu[f"eq_{sha1_short(it['account'] + '|' + it['name'])}"] = bu_id
            cp("")

    # ─── Phase B: interactive creation by keyword hint ──────────────────
    if eq_unassigned:
        cp("═" * 70, C.B)
        cp("Фаза Б. Оборудование без привязки к адресу", C.B)
        cp("═" * 70, C.B)

        by_cat: dict[str, list[dict]] = {}
        for rec in eq_unassigned:
            by_cat.setdefault(_classify_bu_heuristic(rec["name"]), []).append(rec)

        cp(f"Всего без адреса: {len(eq_unassigned)}. Подсказка по эвристике:")
        for cat, items in sorted(by_cat.items(), key=lambda kv: -len(kv[1])):
            cp(f"  • {cat:<35} {len(items):>4} шт.", C.CY)

        cp("\nДействия для каждой эвристической группы:")
        cp("  [Enter] — создать BU с предложенным именем")
        cp("  имя     — создать BU с заданным именем")
        cp("  -       — пропустить (оборудование останется без BU)")
        cp("  =bu_id  — присоединить к ранее созданной BU (введите её id)")

        existing: dict[str, dict] = {b["id"]: b for b in bus}

        for cat, items in sorted(by_cat.items(), key=lambda kv: -len(kv[1])):
            cp(f"\n  Группа: {cat} ({len(items)} объектов)")
            cp("  Примеры: " + " · ".join(it["name"][:40] for it in items[:4]))
            cmd = ask("  Имя BU / '-' / '=bu_id'", default=cat)
            if cmd == "-":
                continue
            target_id: str
            if cmd.startswith("="):
                cand = cmd[1:].strip()
                if cand not in existing:
                    cp(f"    ⚠ BU id «{cand}» не найдена; создаю новую «{cat}»", C.Y)
                    cmd = cat
                else:
                    target_id = cand
                    for it in items:
                        eq_to_bu[f"eq_{sha1_short(it['account'] + '|' + it['name'])}"] = target_id
                    continue
            bu_id = f"bu_{slugify(cmd)}"
            if bu_id not in existing:
                bus.append({
                    "id": bu_id,
                    "name": cmd,
                    "keywords": BU_KEYWORDS.get(cat, []),
                    "address": None,
                    "anchor_cadastral": None,
                    "cadastrals": [],
                    "parent_id": None,
                    "external_ids": {},
                    "source": "user_keyword",
                    "confirmed": True,
                })
                existing[bu_id] = bus[-1]
            target_id = bu_id
            for it in items:
                eq_to_bu[f"eq_{sha1_short(it['account'] + '|' + it['name'])}"] = target_id

    return bus, eq_to_bu


# ════════════════════════════════════════════════════════════════════════════
#  Оборудование + Lease contracts (01.К ↔ 01.03)
# ════════════════════════════════════════════════════════════════════════════
def _lease_key(name: str) -> str:
    """Нормализация имени для сопоставления 01.К ↔ 01.03."""
    s = name.lower()
    s = re.sub(r"\bдог[а-я]*\s*\S*", "", s)
    s = re.sub(r"\s+от\s+\d{2}\.\d{2}\.\d{4}", "", s)
    s = re.sub(r"№\s*\S+", "", s)
    s = re.sub(r":\d{2,8}\b", "", s)            # снимаем кадастровые хвосты
    s = re.sub(r"\d{2}:\d{2}:\d{2,8}:\d{1,8}", "", s)  # и полные КН
    return re.sub(r"\s+", " ", s).strip()


def build_equipment(
    records: list[dict],
    eq_to_bu: dict[str, str],
    hint_to_cad_id: dict[str, str],
) -> tuple[list[dict], list[dict]]:
    """
    Создаёт записи оборудования с **множественной** привязкой:
      links.business_unit_ids[], cadastre_ids[], level_ids[], premises_ids[],
      location_kinds[] (один на каждый cad-таргет).

    location_kind:
      'level'       — оборудование на уровне (этаже) внутри строения
      'premises'    — оборудование закреплено за конкретным помещением
      'land_point'  — на земельном участке (точка / уличное оборудование)
      'land_contour'— по контуру участка (ландшафт)
      'standalone'  — без привязки к недвижимости
    """
    equipment: list[dict] = []
    leases: list[dict] = []
    name_to_eq01_03: dict[str, str] = {}

    for rec in records:
        eq_id = f"eq_{sha1_short(rec['account'] + '|' + rec['name'])}"
        outdoor = _is_outdoor(rec["name"])

        # Множество кадастровых таргетов — все hint'ы, разрешившиеся в КН
        cad_ids: list[str] = []
        for h in rec["cadastral_hints"]:
            cid = hint_to_cad_id.get(h)
            if cid and cid not in cad_ids:
                cad_ids.append(cid)

        if cad_ids and outdoor:
            location_kinds = ["land_point"] * len(cad_ids)
        elif cad_ids:
            location_kinds = ["level"] * len(cad_ids)
        else:
            location_kinds = ["standalone"]

        bu_id = eq_to_bu.get(eq_id)
        equipment.append({
            "id": eq_id,
            "name": rec["name"],
            "account": rec["account"],
            "right_type": rec["right_type"],
            "inv_number_hint": rec["inv_hint"],
            "cadastral_hints": rec["cadastral_hints"],
            "amounts": rec["amounts"],
            "links": {
                "business_unit_ids": [bu_id] if bu_id else [],
                "cadastre_ids":      cad_ids,
                "level_ids":         [],
                "premises_ids":      [],
                "location_kinds":    location_kinds,
            },
            "confirmed": False,
        })
        if rec["account"] == "01.03":
            name_to_eq01_03[_lease_key(rec["name"])] = eq_id

    for rec in records:
        if rec["account"] != "01.К":
            continue
        linked_eq = name_to_eq01_03.get(_lease_key(rec["name"]))
        leases.append({
            "id": f"lc_{sha1_short(rec['name'])}",
            "equipment_ids": [linked_eq] if linked_eq else [],
            "name_reference": rec["name"],
            "cadastral_hint": rec["cadastral_hints"][0] if rec["cadastral_hints"] else None,
            "contract_no": rec["contract_no"],
            "contract_date": rec["contract_date"],
            "rent_quarterly_nu": rec["amounts"]["nu"]["turnover_kt"],
            "linked_account": "01.К",
        })
    return equipment, leases


# ════════════════════════════════════════════════════════════════════════════
#  Авто-привязка оборудования к уровням / помещениям (если детерминировано)
# ════════════════════════════════════════════════════════════════════════════
def auto_link_levels_and_premises(
    equipment: list[dict],
    cadastre_objects: list[dict],
) -> None:
    """
    Если у привязанного кадастрового объекта ровно один уровень — оборудование
    получает level_ids=[этот уровень]. Если объект — Помещение, оборудование
    получает premises_ids=[этот объект]. Множественные привязки сохраняются.
    """
    cad_by_id = {c["id"]: c for c in cadastre_objects}
    for eq in equipment:
        links = eq.get("links", {})
        cad_ids = list(links.get("cadastre_ids") or [])
        new_levels: list[str] = list(links.get("level_ids") or [])
        new_premises: list[str] = list(links.get("premises_ids") or [])
        new_kinds: list[str] = list(links.get("location_kinds") or [])

        for i, cid in enumerate(cad_ids):
            cad = cad_by_id.get(cid)
            if not cad:
                continue
            obj_type = (cad.get("object_type") or "").lower()
            levels = cad.get("levels") or []

            if "помещен" in obj_type:
                if cid not in new_premises:
                    new_premises.append(cid)
                if i < len(new_kinds):
                    new_kinds[i] = "premises"
                else:
                    new_kinds.append("premises")
                continue

            if len(levels) == 1:
                lvl_id = levels[0].get("id")
                if lvl_id and lvl_id not in new_levels:
                    new_levels.append(lvl_id)
                if i < len(new_kinds):
                    new_kinds[i] = "level"

        links["level_ids"] = new_levels
        links["premises_ids"] = new_premises
        links["location_kinds"] = new_kinds
        eq["links"] = links


# ════════════════════════════════════════════════════════════════════════════
#  Cross-link с enrich-output (03_enrich_v11)
# ════════════════════════════════════════════════════════════════════════════
def link_with_enriched(
    enterprise: dict,
    business_units: list[dict],
    cadastre_objects: list[dict],
    extras: dict,
    eq_to_bu: dict[str, str] | None = None,
    equipment: list[dict] | None = None,
) -> None:
    """
    Связывает наши BU с bu::<sha1> из enriched по якорному КН и
    подтягивает ИНН/ОГРН головной компании в enterprise.

    Если найден enrich-ключ — **заменяет** локальный `bu.id` на `bu::<sha1>`
    («поглощение»), чтобы граф 04_nspd_graph_v11 рисовал одну BU вместо
    двух. Также обновляет `eq_to_bu` и `equipment[].links.business_unit_id`.

    Контракт enriched_*.json (из 03_enrich_v11):
      data.business_units:[{Ключ, Наименование, Объект (КН), Бенефициар (ключ)}]
      data.beneficiaries:{<key>: {ИНН, ОГРН, КПП, Полное наименование, ...}}
    """
    enrich_bus = extras.get("business_units", [])
    beneficiaries = extras.get("beneficiaries", {})

    bu_by_cn: dict[str, dict] = {}
    for b in enrich_bus:
        cn = b.get("Объект (КН)") or b.get("объект_кн")
        if cn:
            bu_by_cn[cn] = b

    id_remap: dict[str, str] = {}
    for bu in business_units:
        anchor = bu.get("anchor_cadastral")
        if not anchor or anchor not in bu_by_cn:
            continue
        eb = bu_by_cn[anchor]
        new_id = eb.get("Ключ") or eb.get("ключ")
        if new_id and new_id != bu["id"]:
            id_remap[bu["id"]] = new_id
            bu["id"] = new_id
        bu["external_ids"] = {**bu.get("external_ids", {}), "enrich_bu_key": new_id}
        bu["beneficiary_key"] = eb.get("Бенефициар (ключ)") or eb.get("бенефициар_ключ")

    if id_remap:
        if eq_to_bu is not None:
            for k, v in list(eq_to_bu.items()):
                if v in id_remap:
                    eq_to_bu[k] = id_remap[v]
        if equipment is not None:
            for eq in equipment:
                links = eq.get("links", {}) or {}
                bu_ids = links.get("business_unit_ids") or []
                if any(b in id_remap for b in bu_ids):
                    links["business_unit_ids"] = [id_remap.get(b, b) for b in bu_ids]
                    eq["links"] = links

    # ИНН/ОГРН enterprise: ищем бенефициара с типом «юр. лицо», у которого
    # name_short сильно похоже на enterprise.name_short.
    candidate: dict | None = None
    target = (enterprise.get("name_short") or "").upper()
    for key, ben in beneficiaries.items():
        if not isinstance(ben, dict):
            continue
        full = (ben.get("Полное наименование") or ben.get("Наименование (отображаемое)") or "").upper()
        short = (ben.get("Краткое наименование") or "").upper()
        if target and (target in full or target in short or full.startswith(target[:10])):
            candidate = ben
            break
    if candidate:
        enterprise["inn"]  = candidate.get("ИНН")  or enterprise.get("inn")
        enterprise["ogrn"] = candidate.get("ОГРН") or enterprise.get("ogrn")
        enterprise["kpp"]  = candidate.get("КПП")  or enterprise.get("kpp")
        enterprise["external_ids"] = {
            **enterprise.get("external_ids", {}),
            "enrich_beneficiary_key": next((k for k, v in beneficiaries.items() if v is candidate), None),
        }


# ════════════════════════════════════════════════════════════════════════════
#  Пространственные высоты (русская система: этаж = 3 м)
# ════════════════════════════════════════════════════════════════════════════
LEVEL_HEIGHT_M  = 3.0      # высота полного этажа
ANTRESOL_OFFSET = 2.0      # антресоль = +2 м над верхним этажом
Z_SEMIBASEMENT  = -2.0     # полуподвал / цокольный этаж
Z_TECH_UNDER    = -1.0     # техническое подполье

UNASSIGNED_FOLDER       = "00_Нераспределенные"
UNASSIGNED_FOLDER_LOWER = UNASSIGNED_FOLDER.lower()


def compute_level_z(level: dict, building_above_n: int = 0) -> float | None:
    """
    Z-координата уровня (метры над «уровнем 0», т.е. отметка пола 1 этажа).
    Русская система:
      • подвал K     → -K*3   (подвал 1 = -3, подвал 2 = -6, …)
      • полуподвал   → -2
      • цокольный    → -2  (равен полуподвалу по нашей договорённости)
      • тех.подполье → -1
      • этаж K (надз.)→ (K-1)*3
      • антресоль K  → (K-1)*3 + 2     (антресоль 1-го этажа = 2)
      • антресоль без указания → (N-1)*3 + 2, N = building_above_n
      • мансарда / чердак / эксплуат. кровля → N*3
      • технический этаж (сверху)    → (N+1)*3
    """
    if not isinstance(level, dict):
        return None
    t = (level.get("type") or "").lower()
    label = (level.get("label") or "").lower()
    underground = bool(level.get("underground"))

    if underground:
        if "полуподвал" in t:
            return Z_SEMIBASEMENT
        if "цоколь" in t:
            return Z_SEMIBASEMENT
        if "тех" in t and "подпол" in t:
            return Z_TECH_UNDER
        # Подвал K — пытаемся вытащить порядковый номер
        m = re.search(r"подвал\s*(\d+)|(\d+)\s*подвал", label)
        K = int(next((g for g in (m.groups() if m else ()) if g), "1") or "1") if m else 1
        return -K * LEVEL_HEIGHT_M

    # Надземные
    if "антресол" in t:
        m = re.search(r"антресол\w*\s+(\d+)\s*эта", label)
        if m:
            K = int(m.group(1))
            return (K - 1) * LEVEL_HEIGHT_M + ANTRESOL_OFFSET
        # Антресоль без указания — над последним надземным
        if building_above_n > 0:
            return (building_above_n - 1) * LEVEL_HEIGHT_M + ANTRESOL_OFFSET
        return ANTRESOL_OFFSET

    if "мансарда" in t or "чердак" in t or "эксплуатируемая кровля" in t or "надстро" in t:
        return building_above_n * LEVEL_HEIGHT_M if building_above_n else LEVEL_HEIGHT_M

    if "технич" in t:
        # технический этаж — сверху
        return (building_above_n + 1) * LEVEL_HEIGHT_M if building_above_n else 0.0

    # Обычный этаж: ищем номер K в label («Уровень N. Этаж K»)
    m = re.search(r"этаж\s+(\d+)|\bЭтаж\s+(\d+)", level.get("label") or "")
    if m:
        K = int(next((g for g in m.groups() if g), "1"))
        return (K - 1) * LEVEL_HEIGHT_M

    # Fallback: используем number (глобальный, снизу вверх). Если есть подземные,
    # «number=K» означает K-й уровень от низа, что даёт некорректный z.
    # Поэтому без явных подсказок возвращаем None — пусть остаётся не вычислен.
    return None


def annotate_levels_with_z(cadastre_objects: list[dict]) -> None:
    """Расставляет level.z_meters и cad.height_m / cad.depth_m."""
    for cad in cadastre_objects:
        levels = cad.get("levels") or []
        above_n = sum(1 for lvl in levels if not lvl.get("underground"))
        below_n = sum(1 for lvl in levels if lvl.get("underground"))
        for lvl in levels:
            z = compute_level_z(lvl, building_above_n=above_n)
            if z is not None:
                lvl["z_meters"]      = round(z, 3)
                lvl["height_m"]      = LEVEL_HEIGHT_M  # высота самого уровня
                lvl["top_z_meters"]  = round(z + LEVEL_HEIGHT_M, 3)
        # высота надземной части здания
        if above_n:
            cad["height_m"] = above_n * LEVEL_HEIGHT_M
        if below_n:
            cad["depth_m"]  = below_n * LEVEL_HEIGHT_M
        # объёмная высота для extrude (для KMZ)
        if cad.get("height_m") and cad.get("_geometry"):
            geom = cad["_geometry"]
            geom["bottom_z_meters"] = -(cad.get("depth_m") or 0.0)
            geom["top_z_meters"]    = cad["height_m"]
            geom["extrude"] = True


# ════════════════════════════════════════════════════════════════════════════
#  EXIF (Pillow, опционально) — поля как у viewer-а: lat/lon/alt/bearing/date
# ════════════════════════════════════════════════════════════════════════════
try:
    from PIL import Image, ExifTags  # type: ignore
    _PIL_OK = True
except Exception:
    _PIL_OK = False
    Image = None  # type: ignore
    ExifTags = None  # type: ignore


def _dms_to_decimal(dms, ref) -> float | None:
    """Преобразует EXIF DMS (3 рацион. числа) в десятичную долю."""
    if not dms:
        return None
    try:
        d, m, s = dms
        val = float(d) + float(m) / 60.0 + float(s) / 3600.0
    except Exception:
        try:
            val = float(dms[0]) + float(dms[1]) / 60.0 + float(dms[2]) / 3600.0
        except Exception:
            return None
    if ref in ("S", "W", b"S", b"W"):
        val = -val
    return val


def read_exif(path: Path) -> dict:
    """
    Читает GPS / дату / камеру из JPEG/HEIC через Pillow.
    Возвращает: {gps_lat, gps_lon, gps_alt, gps_bearing, datetime_taken,
                 camera_make, camera_model}. Отсутствующие поля = None.
    Если Pillow не установлен или файл нечитаем — пустой dict.
    """
    if not _PIL_OK:
        return {}
    try:
        img = Image.open(path)
        exif = img.getexif()
    except Exception:
        return {}
    if not exif:
        return {}

    by_name: dict[str, Any] = {}
    for tag_id, value in exif.items():
        name = ExifTags.TAGS.get(tag_id, tag_id)
        by_name[name] = value

    gps_info = {}
    if ExifTags.IFD and "GPSInfo" in by_name:
        try:
            gps_ifd = exif.get_ifd(ExifTags.IFD.GPSInfo)
        except Exception:
            gps_ifd = by_name.get("GPSInfo") or {}
        if isinstance(gps_ifd, dict):
            for k, v in gps_ifd.items():
                tag_name = ExifTags.GPSTAGS.get(k, k)
                gps_info[tag_name] = v

    lat = _dms_to_decimal(gps_info.get("GPSLatitude"), gps_info.get("GPSLatitudeRef"))
    lon = _dms_to_decimal(gps_info.get("GPSLongitude"), gps_info.get("GPSLongitudeRef"))
    alt = gps_info.get("GPSAltitude")
    try:
        alt = float(alt) if alt is not None else None
        if alt is not None and gps_info.get("GPSAltitudeRef") in (1, b"\x01"):
            alt = -alt
    except Exception:
        alt = None
    bearing = gps_info.get("GPSImgDirection")
    try:
        bearing = float(bearing) if bearing is not None else None
    except Exception:
        bearing = None
    date = by_name.get("DateTimeOriginal") or by_name.get("DateTime")
    if isinstance(date, bytes):
        date = date.decode("utf-8", "ignore")

    return {
        "gps_lat": lat,
        "gps_lon": lon,
        "gps_alt": alt,
        "gps_bearing": bearing,
        "datetime_taken": str(date) if date else None,
        "camera_make": str(by_name.get("Make") or "") or None,
        "camera_model": str(by_name.get("Model") or "") or None,
    }


# ════════════════════════════════════════════════════════════════════════════
#  Сканер папок с фото → автоматическая привязка к объектам структуры
# ════════════════════════════════════════════════════════════════════════════
# Маска КН в имени папки / файла: «:» → «_», «/» → «__»
# Границы — не цифры, чтобы не захватывать середину длинного числа.
CN_MASK_RE      = re.compile(r"(?<!\d)(\d{2})_(\d{2})_(\d{2,8})_(\d{1,8})(?:__(\d+))?(?!\d)")
INV_MASK_RE     = re.compile(r"(?<!\d)01[_\.]\d{2,5}(?:[_\.]\d{2,5})?(?!\d)")
PHOTO_EXTENSIONS = (".jpg", ".jpeg", ".png", ".heic", ".tif", ".tiff", ".bmp", ".webp")


def _cn_from_mask(s: str) -> list[str]:
    """Восстанавливает КН (полный и part-of) из строки с подчёркиваниями."""
    out: list[str] = []
    for m in CN_MASK_RE.finditer(s):
        cn = f"{m.group(1)}:{m.group(2)}:{m.group(3)}:{m.group(4)}"
        if m.group(5):
            cn = f"{cn}/{m.group(5)}"
        if cn not in out:
            out.append(cn)
    return out


def _normalize_folder_name(name: str) -> str:
    """Для эвристического сопоставления BU/equipment по имени папки."""
    s = name.lower()
    s = re.sub(r"[\s\-\.]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def scan_photo_folder(
    root: Path,
    cadastre_objects: list[dict],
    business_units: list[dict],
    equipment: list[dict],
) -> list[dict]:
    """
    Рекурсивно обходит корневую папку с фото и привязывает каждый файл к
    объектам структуры. Возвращает массив photos[].

    Каждое фото:
      • read_exif → gps_lat/gps_lon/gps_alt/bearing/datetime/camera;
      • linked: cadastre/equipment/business_unit/level/premises (массивы);
      • z_meters: высота над уровнем 0 (отметка пола 1 этажа), вычисляется
        из level.z_meters (если привязан к уровню), либо из EXIF.gps_alt;
      • is_unassigned=True, если файл лежит в папке 00_Нераспределенные —
        в этом случае z_meters остаётся None (без обработки высотами).
    """
    if not root.exists() or not root.is_dir():
        return []
    cad_by_cn = {c["cadastral_number"]: c for c in cadastre_objects}
    cad_by_id = {c["id"]: c for c in cadastre_objects}
    eq_by_inv = {eq["inv_number_hint"]: eq for eq in equipment if eq.get("inv_number_hint")}
    bu_by_slug = {slugify(bu["name"]): bu for bu in business_units}
    photos: list[dict] = []

    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in PHOTO_EXTENSIONS:
            continue
        rel_parts = path.relative_to(root).parts
        folder_chain = list(rel_parts[:-1])
        combined = " / ".join(folder_chain + [path.name])

        # Папка «00_Нераспределенные» — только в корне, по уточнению пользователя
        is_unassigned = bool(folder_chain) and folder_chain[0].lower() == UNASSIGNED_FOLDER_LOWER

        linked = {
            "cadastre_ids": [], "business_unit_ids": [],
            "equipment_ids": [], "level_ids": [], "premises_ids": [],
        }
        match_kinds: list[str] = []

        if not is_unassigned:
            for cn in _cn_from_mask(combined):
                cad = cad_by_cn.get(cn)
                if cad and cad["id"] not in linked["cadastre_ids"]:
                    linked["cadastre_ids"].append(cad["id"])
                    match_kinds.append("cn_mask")
                    if "помещен" in (cad.get("object_type") or "").lower():
                        linked["premises_ids"].append(cad["id"])

            for inv_m in INV_MASK_RE.finditer(combined):
                inv = inv_m.group(0)
                eq = eq_by_inv.get(inv)
                if eq and eq["id"] not in linked["equipment_ids"]:
                    linked["equipment_ids"].append(eq["id"])
                    match_kinds.append("inv_hint")

            for folder in folder_chain:
                slug = _normalize_folder_name(folder)
                for bu_slug, bu in bu_by_slug.items():
                    if bu_slug and bu_slug in slug and bu["id"] not in linked["business_unit_ids"]:
                        linked["business_unit_ids"].append(bu["id"])
                        match_kinds.append("bu_name")

            if not any(linked.values()):
                for folder in folder_chain:
                    fn = folder.lower()
                    for cad in cadastre_objects:
                        addr = (cad.get("address") or "").lower()
                        if not addr:
                            continue
                        tokens = {t for t in re.split(r"[\s,]+", addr) if len(t) >= 4}
                        overlap = sum(1 for t in tokens if t in fn)
                        if overlap >= 2 and cad["id"] not in linked["cadastre_ids"]:
                            linked["cadastre_ids"].append(cad["id"])
                            match_kinds.append("address_overlap")
                            break

        exif = read_exif(path)
        if exif:
            match_kinds.append("exif")

        # z_meters: приоритет — этаж объекта; иначе — EXIF altitude (если есть);
        # если is_unassigned → высот не вычисляем (только сырой EXIF сохраняем).
        z_meters: float | None = None
        z_source: str | None = None
        if not is_unassigned:
            for cad_id in linked["cadastre_ids"]:
                cad = cad_by_id.get(cad_id)
                levels = (cad or {}).get("levels") or []
                if len(levels) == 1 and levels[0].get("z_meters") is not None:
                    z_meters = levels[0]["z_meters"]
                    z_source = "level"
                    break
            if z_meters is None and isinstance(exif.get("gps_alt"), (int, float)):
                z_meters = float(exif["gps_alt"])
                z_source = "exif_gps_alt"

        photo_id = "ph_" + sha1_short(str(path).replace("\\", "/"))
        photos.append({
            "id": photo_id,
            "path": str(path),
            "folder_chain": folder_chain,
            "is_unassigned": is_unassigned,
            "linked": linked,
            "match_kinds": sorted(set(match_kinds)),
            "exif": exif,
            "z_meters": z_meters,
            "z_source": z_source,
        })
    return photos


# ════════════════════════════════════════════════════════════════════════════
#  Генератор структуры папок на диске для размещения фотографий
# ════════════════════════════════════════════════════════════════════════════
# Шаблон: Класс_КН_Параметр (только подчёркивания)
OBJECT_CLASS_FOR_TYPE: dict[str, str] = {
    "Земельный участок":                       "Земельный_участок",
    "Единое землепользование":                 "Земельный_участок",
    "Здание":                                  "Здание",
    "Сооружение":                              "Сооружение",
    "Помещение":                               "Помещение",
    "Объект незавершенного строительства":     "ОНС",
}

OBJECT_CLASSES_WITH_PLAN = {"Здание", "Сооружение", "ОНС"}


def _cn_to_mask(cn: str | None) -> str:
    if not cn:
        return ""
    return cn.replace(":", "_").replace("/", "__")


def _area_param(cad: dict) -> str:
    """Параметр: значение площади / длины для имени папки."""
    area = cad.get("area")
    if area in (None, "", "—", "-"):
        return ""
    # area может быть числом или строкой «868 кв. м», «727,7»
    if isinstance(area, (int, float)):
        return f"{area}_кв.м"
    s = str(area)
    # Удаляем единицы и пробелы, унифицируем разделитель
    m = re.search(r"(\d+[\.,]?\d*)\s*([а-я\.]+)?", s)
    if not m:
        return ""
    num = m.group(1).replace(",", ".")
    unit_raw = (m.group(2) or "").lower()
    if "пог" in unit_raw or "п.м" in unit_raw:
        unit = "п.м"
    elif "куб" in unit_raw:
        unit = "куб.м"
    elif unit_raw == "м" or unit_raw == "м." or "метр" in unit_raw:
        unit = "м"
    else:
        unit = "кв.м"
    return f"{num}_{unit}"


def _safe_folder_segment(s: str) -> str:
    """Очищает имя папки от запрещённых в Windows символов."""
    if not s:
        return ""
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("._ ")


def _folder_name_cadastre(cad: dict) -> str:
    cls = OBJECT_CLASS_FOR_TYPE.get(cad.get("object_type") or "", "Объект")
    cn_mask = _cn_to_mask(cad.get("cadastral_number"))
    param = _area_param(cad)
    parts = [cls, cn_mask]
    if param:
        parts.append(param)
    return _safe_folder_segment("_".join(p for p in parts if p))


def _folder_name_business_unit(bu: dict) -> str:
    parts = ["Бизнес-единица", slugify(bu.get("name") or bu.get("id") or "")]
    return _safe_folder_segment("_".join(parts))


def _folder_name_equipment(eq: dict) -> str:
    parts = ["Оборудование", eq["id"]]
    if eq.get("inv_number_hint"):
        parts.append(eq["inv_number_hint"])
    short = slugify((eq.get("name") or "")[:30])
    if short:
        parts.append(short)
    return _safe_folder_segment("_".join(parts))


def generate_folder_structure(
    root: Path,
    cadastre_objects: list[dict],
    business_units: list[dict],
    equipment: list[dict],
    dry_run: bool = False,
) -> dict[str, list[str]]:
    """
    Создаёт на диске структуру папок для размещения фотографий.

    Layout:
      <root>/
        01_Земельные_участки/Земельный_участок_<КН>_<S>_кв.м/
        02_Здания/Здание_<КН>_<S>_кв.м/План_объекта/
        03_Сооружения/Сооружение_<КН>_<S>_<ед>/План_объекта/
        04_Помещения/Помещение_<КН>_<S>_кв.м/
        05_ОНС/ОНС_<КН>/План_объекта/
        06_Бизнес-единицы/Бизнес-единица_<slug>/
        07_Оборудование/Оборудование_<id>_<inv>_<имя>/

    Возвращает {"created": [...], "skipped": [...]}.
    Идемпотентно: повторный запуск не пересоздаёт существующие папки.
    """
    created: list[str] = []
    skipped: list[str] = []
    if not dry_run:
        root.mkdir(parents=True, exist_ok=True)

    # «00_Нераспределенные» — корневая папка для фото, не подвергаемых
    # обработке высотами (отдельная категория в начале списка).
    nonassigned_path = root / UNASSIGNED_FOLDER
    if nonassigned_path.exists():
        skipped.append(str(nonassigned_path))
    else:
        if not dry_run:
            nonassigned_path.mkdir(parents=True, exist_ok=True)
        created.append(str(nonassigned_path))

    sections: list[tuple[str, str, list[dict], callable]] = [
        ("01_Земельные_участки", "Земельный участок",
         [c for c in cadastre_objects if "земельн" in (c.get("object_type") or "").lower()],
         _folder_name_cadastre),
        ("02_Здания", "Здание",
         [c for c in cadastre_objects if c.get("object_type") == "Здание"],
         _folder_name_cadastre),
        ("03_Сооружения", "Сооружение",
         [c for c in cadastre_objects if c.get("object_type") == "Сооружение"],
         _folder_name_cadastre),
        ("04_Помещения", "Помещение",
         [c for c in cadastre_objects if "помещен" in (c.get("object_type") or "").lower()],
         _folder_name_cadastre),
        ("05_ОНС", "ОНС",
         [c for c in cadastre_objects if c.get("object_type") == "Объект незавершенного строительства"],
         _folder_name_cadastre),
        ("06_Бизнес-единицы", "Бизнес-единица", business_units, _folder_name_business_unit),
        ("07_Оборудование", "Оборудование", equipment, _folder_name_equipment),
    ]

    needs_plan = OBJECT_CLASSES_WITH_PLAN  # Здание, Сооружение, ОНС
    type_class_map = OBJECT_CLASS_FOR_TYPE

    for section_dir, _class, items, namer in sections:
        if not items:
            continue
        section_path = root / section_dir
        if not dry_run:
            section_path.mkdir(parents=True, exist_ok=True)
        for item in items:
            fname = namer(item)
            if not fname:
                continue
            fpath = section_path / fname
            if fpath.exists():
                skipped.append(str(fpath))
            else:
                if not dry_run:
                    fpath.mkdir(parents=True, exist_ok=True)
                created.append(str(fpath))
            # Для здания / сооружения / ОНС — подпапка «План_объекта»
            cls_short = type_class_map.get(item.get("object_type") or "", "")
            if cls_short in needs_plan:
                plan = fpath / "План_объекта"
                if plan.exists():
                    skipped.append(str(plan))
                else:
                    if not dry_run:
                        plan.mkdir(parents=True, exist_ok=True)
                    created.append(str(plan))

    return {"created": created, "skipped": skipped}


# ════════════════════════════════════════════════════════════════════════════
#  Идемпотентный merge
# ════════════════════════════════════════════════════════════════════════════
USER_EDITABLE_FIELDS = ("name", "links", "address", "confirmed", "photo_paths")


def merge_preserve_confirmed(old: dict, new: dict) -> dict:
    """Сохраняет confirmed=true записи и user-editable поля из старой структуры."""
    if not old:
        return new

    def index(lst: list[dict]) -> dict[str, dict]:
        return {x["id"]: x for x in lst if isinstance(x, dict) and "id" in x}

    for section in ("business_units", "cadastre_objects", "premises",
                    "equipment", "lease_contracts", "photos"):
        old_idx = index(old.get(section, []))
        merged: list[dict] = []
        seen: set[str] = set()
        for nv in new.get(section, []):
            ov = old_idx.get(nv["id"])
            if ov and ov.get("confirmed") is True:
                merged.append(ov)
            elif ov:
                combined = {**nv, **{
                    k: v for k, v in ov.items()
                    if k in USER_EDITABLE_FIELDS and v not in (None, "", [], {})
                }}
                merged.append(combined)
            else:
                merged.append(nv)
            seen.add(nv["id"])
        for oid, ov in old_idx.items():
            if oid not in seen and ov.get("confirmed") is True:
                merged.append(ov)
        new[section] = merged
    return new


# ════════════════════════════════════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════════════════════════════════════
def _intro() -> None:
    cp("\n" + "═" * 70, C.B)
    cp("  ОСВ → ИЕРАРХИЯ JSON: предприятие · BU · кадастр · оборудование", C.B)
    cp("═" * 70, C.B)
    cp("Что потребуется на вход:")
    cp("  1) Файл ОСВ счёта 01 в формате .xlsx (обязательно).")
    cp("  2) Предыдущий structure_*.json (опционально, для обогащения).")
    cp("  3) Каталог JSON-выгрузок NSPD/ЕГРН (опционально, для адресов).")
    cp("  4) Корневая папка с фотографиями (опционально).")
    cp("  В конце спросим: создать ли структуру папок для фото на диске.\n")


def main() -> None:
    _intro()

    osv_path_str = ask("Путь к .xlsx ОСВ счёта 01")
    if not osv_path_str:
        cp("Отмена.", C.R)
        return
    osv_path = Path(osv_path_str)
    if not osv_path.exists() or osv_path.suffix.lower() != ".xlsx":
        cp(f"Файл не найден или не .xlsx: {osv_path}", C.R)
        return

    prev_data: dict | None = None
    if prev_path_str := ask("Путь к structure_*.json (Enter — пропустить)"):
        p = Path(prev_path_str)
        if p.exists():
            prev_data = load_json_safe(p)
            cp(f"  ✓ Загружено: {p.name}", C.G)
        else:
            cp(f"  ⚠ Файл не найден, продолжаю без merge: {p}", C.Y)

    nspd_items: list[dict] = []
    enriched_extras: dict[str, Any] = {"beneficiaries": {}, "business_units": []}
    if nspd_dir_str := ask("Каталог JSON-выгрузок NSPD/ЕГРН/enriched (Enter — пропустить)"):
        d = Path(nspd_dir_str)
        nspd_items = load_nspd_objects(d)
        enriched_extras = load_enriched_extras(d)
        cp(f"  ✓ Объектов: {len(nspd_items)} · бенефициаров: "
           f"{len(enriched_extras['beneficiaries'])} · BU (enrich): "
           f"{len(enriched_extras['business_units'])}", C.G)

    photo_dir_str = ask("Корневая папка с фотографиями (Enter — пропустить)")

    # ── Парсинг ─────────────────────────────────────────────────────────
    cp("\n⏳ Парсинг ОСВ…", C.CY)
    enterprise, records = parse_osv(osv_path)
    cp(f"  предприятие: {enterprise['opf']} «{enterprise['name_short']}»", C.G)
    cp(f"  период:      {enterprise['period']['label']}", C.G)
    cp(f"  ОС-записей:  {len(records)}", C.G)

    by_acc: dict[str, int] = {}
    for r in records:
        by_acc[r["account"]] = by_acc.get(r["account"], 0) + 1
    for a, n in sorted(by_acc.items()):
        right = (ACCOUNT_MAP.get(a, {}) or {}).get("right") or "—"
        cp(f"    {a:<6} ({right:<14}): {n}", C.CY)

    # ── Кадастр ─────────────────────────────────────────────────────────
    cp("\n⏳ Сборка кадастровых объектов…", C.CY)
    cadastre_objects, hint_to_cad_id = build_cadastre_objects(records, nspd_items)
    for cad in cadastre_objects:
        if cad.get("_raw_info"):
            cad["levels"] = compute_levels_for_cadastre(cad)
            cad.pop("_raw_info", None)
    annotate_levels_with_z(cadastre_objects)
    with_height = sum(1 for c in cadastre_objects if c.get("height_m"))
    with_lvl_z  = sum(1 for c in cadastre_objects for lvl in (c.get("levels") or []) if lvl.get("z_meters") is not None)
    cp(f"  кадастровых объектов: {len(cadastre_objects)} (с высотой: {with_height}, уровней с z: {with_lvl_z})", C.G)

    # ── Бизнес-единицы ──────────────────────────────────────────────────
    cp("\n⏳ Построение бизнес-единиц (приоритет — по адресам)…", C.CY)
    business_units, eq_to_bu = build_business_units(records, cadastre_objects, hint_to_cad_id)
    cp(f"  бизнес-единиц создано: {len(business_units)}", C.G)

    # ── Оборудование + аренда ───────────────────────────────────────────
    cp("\n⏳ Сборка оборудования и договоров аренды…", C.CY)
    equipment, lease_contracts = build_equipment(records, eq_to_bu, hint_to_cad_id)

    # ── Авто-привязка к уровням и помещениям (множественная) ───────────
    auto_link_levels_and_premises(equipment, cadastre_objects)
    auto_lvl = sum(1 for e in equipment if e["links"].get("level_ids"))
    auto_prem = sum(1 for e in equipment if e["links"].get("premises_ids"))
    cp(f"  авто-уровней: {auto_lvl} · авто-помещений: {auto_prem}", C.G)

    # ── Высота оборудования по уровню (z_meters) ───────────────────────
    lvl_z_by_id: dict[str, float] = {}
    for cad in cadastre_objects:
        for lvl in cad.get("levels", []) or []:
            if lvl.get("id") and lvl.get("z_meters") is not None:
                lvl_z_by_id[lvl["id"]] = lvl["z_meters"]
    eq_with_z = 0
    for eq in equipment:
        lvls = eq["links"].get("level_ids") or []
        z_vals = [lvl_z_by_id[l] for l in lvls if l in lvl_z_by_id]
        if z_vals:
            eq["z_meters"] = min(z_vals)
            eq["z_meters_max"] = max(z_vals) + LEVEL_HEIGHT_M
            eq_with_z += 1
        else:
            eq["z_meters"] = None
    cp(f"  оборудования с высотой (z): {eq_with_z}", C.G)

    # ── Cross-link с enrich-output (поглощение BU + ИНН/ОГРН) ──────────
    if enriched_extras["business_units"] or enriched_extras["beneficiaries"]:
        link_with_enriched(enterprise, business_units, cadastre_objects,
                            enriched_extras, eq_to_bu=eq_to_bu, equipment=equipment)
        cp(f"  cross-link с enrich: ИНН={enterprise.get('inn')} · "
           f"ОГРН={enterprise.get('ogrn')}", C.G)
    cp(f"  оборудование:  {len(equipment)}", C.G)
    cp(f"  lease записей: {len(lease_contracts)}", C.G)

    # ── Сканирование папки с фото (опц.) ───────────────────────────────
    photos: list[dict] = []
    if photo_dir_str:
        pd = Path(photo_dir_str)
        if pd.exists() and pd.is_dir():
            cp(f"\n⏳ Сканирование папки с фото: {pd}…", C.CY)
            photos = scan_photo_folder(pd, cadastre_objects, business_units, equipment)
            matched = sum(1 for p in photos if any(p["linked"].values()))
            cp(f"  файлов: {len(photos)} · с привязкой: {matched}", C.G)
        else:
            cp(f"  ⚠ Папка не найдена: {pd}", C.Y)

    # ── Сборка финального документа ─────────────────────────────────────
    structure: dict[str, Any] = {
        "meta": {
            "schema_version": "1.0",
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "source_files": [str(osv_path)],
            "run_id": sha1_short(datetime.now().isoformat() + str(osv_path), 12),
        },
        "enterprise": enterprise,
        "dictionaries": {
            "level_types":  LEVEL_TYPES,
            "opf_map":      OPF_MAP,
            "account_map":  ACCOUNT_MAP,
            "bu_keywords":  BU_KEYWORDS,
        },
        "business_units": business_units,
        "cadastre_objects": cadastre_objects,
        "premises": (prev_data or {}).get("premises", []),
        "equipment": equipment,
        "lease_contracts": lease_contracts,
        "photos": photos,
    }

    if prev_data:
        cp("\n⏳ Идемпотентный merge с предыдущим JSON…", C.CY)
        structure = merge_preserve_confirmed(prev_data, structure)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = osv_path.parent / f"structure_{enterprise['slug']}_{ts}.json"
    out.write_text(json.dumps(structure, ensure_ascii=False, indent=4), encoding="utf-8")

    cp("\n" + "═" * 70, C.B)
    cp("ГОТОВО", C.B + C.G)
    cp("═" * 70, C.B)
    cp(f"  Файл: {out}", C.CY)
    cp(f"  Записей: ОС={len(equipment)} · BU={len(business_units)} · "
       f"кадастр={len(cadastre_objects)} · аренда={len(lease_contracts)}", C.G)
    if photos:
        cp(f"  Фотографий: {len(photos)}", C.G)

    # ── Опция: сгенерировать структуру папок для фото ──────────────────
    if ask_yn("\nСгенерировать структуру папок на диске для размещения фото?",
              default=False):
        gen_root_str = ask("  Корневая папка для генерации",
                            default=str(osv_path.parent / "ФОТО"))
        gen_root = Path(gen_root_str)
        cp(f"\n⏳ Создание структуры папок в {gen_root}…", C.CY)
        result = generate_folder_structure(
            gen_root, structure["cadastre_objects"],
            structure["business_units"], structure["equipment"],
        )
        cp(f"  создано: {len(result['created'])} · пропущено (уже есть): "
           f"{len(result['skipped'])}", C.G)
        if result["created"][:3]:
            cp("  примеры:", C.CY)
            for p in result["created"][:3]:
                cp(f"    {p}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        cp("\n\nПрервано пользователем (Ctrl+C).", C.R)
        sys.exit(1)
    except Exception as e:
        cp(f"\n✗ Непредвиденная ошибка: {e}", C.R)
        import traceback
        traceback.print_exc()
        sys.exit(1)
