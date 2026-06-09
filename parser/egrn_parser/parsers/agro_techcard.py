"""
egrn_parser/parsers/agro_techcard.py — парсер техкарты виноградника (ADR-006, трек A).

ТОЛЬКО виноградники: при отсутствии маркеров винограда `parse_workbook` вернёт
`is_vineyard=False` и пустые записи (работы по другим культурам не обрабатываются;
структуру листов можно переиспользовать). Источник §6 — `source='techcard'`.

Листы (3 типа структур):
  • смета-операции (Посадка/Шпалера/Уходные): код | работа | ед | стоим/ед | <год> | ИТОГО;
  • СЗР+удобрения: № | препарат | цена | … | расход на 1 га;
  • плодоносящие (проверка): № | работа | ед | объём | цена | сумма | срок.

Маппинг (ADR-006): meta(площадь/саженцы) → agro_parcel; виноград(perennial) →
agro_crop_cycle(plan); операции/СЗР → agro_event (operation|treatment), показатели
в JSON attrs. События валидируются `agro_event_profiles.validate_event_attrs`.
"""
from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
except ImportError:                      # pragma: no cover
    openpyxl = None

from egrn_parser.parsers import agro_event_profiles as _profiles

_VINE_MARKERS = ("виноград", "саженц", "шпалер")


def _num(v: Any) -> Optional[float]:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _year(v: Any) -> Optional[int]:
    m = re.search(r"(20\d\d)", str(v or ""))
    return int(m.group(1)) if m else None


def _cell(ws, r: int, c: int) -> Any:
    return ws.cell(r, c).value


# ── классификация листа ──────────────────────────────────────────────────────
def _sheet_kind(ws) -> str:
    title = (ws.title or "").lower()
    head = " ".join(str(_cell(ws, r, c) or "")
                    for r in range(1, 14) for c in range(1, 6)).lower()
    if "средства защиты" in title or "средства защиты" in head:
        return "szr"
    if "наименование работ" in head and "объ" in head:
        return "works"
    return "estimate"


# ── парсеры листов ───────────────────────────────────────────────────────────
def _parse_estimate(ws) -> list[dict[str, Any]]:
    """Смета: операция = строка с названием И (ед. изм. ИЛИ стоим/ед).

    Колонки: код=2 · работа=3 · ед=4 · стоим/ед=5 · сумма года=6 · ИТОГО=7 · примеч=8.
    Год берётся из шапки (ячейка col6 с '20NN год'). Секции-подзаголовки без ед.изм.
    пропускаются."""
    year = None
    for r in range(1, min(ws.max_row, 6) + 1):
        y = _year(_cell(ws, r, 6))
        if y:
            year = y
            break
    ops = []
    for r in range(1, ws.max_row + 1):
        name = _cell(ws, r, 3)
        if not name or not str(name).strip():
            continue
        nm = str(name).strip()
        if nm.upper().startswith("ИТОГО"):
            continue
        unit = _cell(ws, r, 4)
        unit_cost = _num(_cell(ws, r, 5))
        if unit_cost is None:                    # секции/подзаголовки/описания без цены/ед
            continue
        ops.append({"code": (str(_cell(ws, r, 2)).strip() if _cell(ws, r, 2) else None),
                    "name": nm, "unit": (str(unit).strip() if unit else None),
                    "unit_cost": unit_cost, "total": _num(_cell(ws, r, 7)),
                    "year": year,
                    "note": (str(_cell(ws, r, 8)).strip() if _cell(ws, r, 8) else None),
                    "phase": ws.title.strip()})
    return ops


def _parse_works(ws) -> list[dict[str, Any]]:
    """Плодоносящие (проверка): № | работа | ед | объём | цена | сумма | срок."""
    year = None
    ops = []
    for r in range(1, ws.max_row + 1):
        y = _year(_cell(ws, r, 7))
        if y and not year:
            year = y
        no = _cell(ws, r, 1)
        name = _cell(ws, r, 2)
        if not isinstance(no, (int, float)) or isinstance(no, bool):
            continue
        if not name or not str(name).strip():
            continue
        ops.append({"code": str(int(no)), "name": str(name).strip(),
                    "unit": (str(_cell(ws, r, 3)).strip() if _cell(ws, r, 3) else None),
                    "qty": _num(_cell(ws, r, 4)), "unit_cost": _num(_cell(ws, r, 5)),
                    "total": _num(_cell(ws, r, 6)), "year": year,
                    "phase": ws.title.strip()})
    return ops


def _parse_szr(ws) -> list[dict[str, Any]]:
    """СЗР + удобрения: № | препарат | цена | … | расход на 1 га.

    Блок пестицидов до строки-заголовка «Удобрения…», далее — удобрения."""
    subs = []
    kind = "pesticide"
    for r in range(1, ws.max_row + 1):
        no = _cell(ws, r, 1)
        name = _cell(ws, r, 2)
        # Заголовок секции «Удобрения…» (в col1 или col2, без номера) → переключение блока.
        marker = " ".join(str(_cell(ws, r, c) or "") for c in (1, 2)).lower()
        if "удобрен" in marker and not isinstance(no, (int, float)):
            kind = "fertilizer"
            continue
        if not isinstance(no, (int, float)) or isinstance(no, bool):
            continue
        if not name or not str(name).strip():
            continue
        nm = str(name).strip()
        if nm.upper().startswith("ИТОГО") or nm.lower().startswith("на 1 га"):
            continue
        subs.append({"name": nm, "kind": kind, "price": _num(_cell(ws, r, 3)),
                     "rate_per_ha": _num(_cell(ws, r, 5))})
    return subs


def _extract_meta(wb) -> dict[str, Any]:
    """Площадь закладки (га) и число саженцев из шапок листов (best-effort)."""
    meta: dict[str, Any] = {"area_ha": None, "saplings": None}
    area_cands = []
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 5) + 1):
            for c in range(1, min(ws.max_column, 9)):
                n = _num(_cell(ws, r, c))
                if n is None:
                    continue
                nxt = str(_cell(ws, r, c + 1) or "").lower()
                note = str(_cell(ws, r, c + 2) or "").lower()
                if nxt.strip() == "га":
                    area_cands.append((n, "закладк" in note))
                if "саженц" in nxt and meta["saplings"] is None:
                    meta["saplings"] = int(n)
    if area_cands:
        planting = [a for a, is_p in area_cands if is_p]
        meta["area_ha"] = planting[0] if planting else max(a for a, _ in area_cands)
    return meta


# ── публичный API ────────────────────────────────────────────────────────────
def parse_workbook(path: str | Path) -> dict[str, Any]:
    """Разобрать xlsx техкарты. Не-виноград → is_vineyard=False, пустые секции."""
    if openpyxl is None:                 # pragma: no cover
        raise RuntimeError("openpyxl не установлен")
    wb = openpyxl.load_workbook(str(path), data_only=True)
    blob = " ".join((ws.title or "") for ws in wb.worksheets).lower()
    blob += " ".join(str(_cell(ws, r, c) or "")
                     for ws in wb.worksheets for r in range(1, 8)
                     for c in range(1, 6)).lower()
    is_vineyard = any(m in blob for m in _VINE_MARKERS)
    out: dict[str, Any] = {"is_vineyard": is_vineyard,
                           "crop": "виноград" if is_vineyard else None,
                           "meta": {}, "operations": [], "substances": []}
    if not is_vineyard:
        return out
    out["meta"] = _extract_meta(wb)
    for ws in wb.worksheets:
        kind = _sheet_kind(ws)
        if kind == "szr":
            out["substances"].extend(_parse_szr(ws))
        elif kind == "works":
            out["operations"].extend(_parse_works(ws))
        else:
            out["operations"].extend(_parse_estimate(ws))
    return out


def to_agro_records(parsed: dict[str, Any], *, parcel_code: str = "виноградник",
                    confidence: float = 0.7) -> dict[str, Any]:
    """parse_workbook → {parcel, cycle, events[]} (ADR-006). Пусто для не-винограда."""
    if not parsed.get("is_vineyard"):
        return {"parcel": None, "cycle": None, "events": []}
    ops = parsed["operations"]
    years = [o["year"] for o in ops if o.get("year")]
    season = max(years) if years else None
    src = {"source": "techcard", "confidence": confidence}
    parcel = {"parcel_code": parcel_code, "season_year": season,
              "area_ha": parsed["meta"].get("area_ha"), **src}
    planting_year = min(years) if years else None
    cycle = {"crop": "виноград", "cycle_kind": "perennial",
             "planting_year": planting_year,                 # для отчётов
             "sow_date": (str(planting_year) if planting_year else None),  # закладка (ADR §I)
             "season_year": season, "crop_status": "plan", **src}
    events = []
    for o in ops:
        # Строки сметы — обобщённые операции (event_type='operation'). Реальные
        # обработки СЗР с действующими веществами — ниже, из листа СЗР.
        attrs = {"work": o["name"], "code": o.get("code"), "phase": o.get("phase"),
                 "unit": o.get("unit"), "unit_cost": o.get("unit_cost"),
                 "qty": o.get("qty"), "total": o.get("total")}
        events.append({"event_type": "operation", "season_year": o.get("year") or season,
                       "attrs": {k: v for k, v in attrs.items() if v is not None}, **src})
    for s in parsed["substances"]:
        if s["kind"] == "pesticide":
            events.append({"event_type": "treatment", "season_year": season, **src,
                           "attrs": {"kind": "СЗР (план)", "preparation": s["name"],
                                     "active_substances": [{"name": s["name"],
                                         "rate": s["rate_per_ha"], "unit": "на га"}]}})
        else:
            events.append({"event_type": "operation", "season_year": season, **src,
                           "attrs": {"kind": "удобрение (план)", "name": s["name"],
                                     "rate_per_ha": s["rate_per_ha"]}})
    return {"parcel": parcel, "cycle": cycle, "events": events}


def ingest(conn: sqlite3.Connection, path: str | Path, *,
           parcel_code: str = "виноградник") -> dict[str, Any]:
    """Разобрать техкарту и записать в агро-слой (agro_parcel/agro_crop_cycle/
    agro_event). Требует применённой миграции 0005. Невалидные по профилю (D)
    события всё равно пишутся (это план), их число — в результате."""
    parsed = parse_workbook(path)
    rec = to_agro_records(parsed, parcel_code=parcel_code)
    if not rec["parcel"]:
        return {"is_vineyard": False, "written": {"parcel": 0, "events": 0}, "invalid": []}
    p = rec["parcel"]
    parcel_id = conn.execute(
        "INSERT INTO agro_parcel(parcel_code, season_year, area_ha, source, confidence) "
        "VALUES(?,?,?,?,?)",
        (p["parcel_code"], p["season_year"], p["area_ha"], p["source"], p["confidence"])
    ).lastrowid
    c = rec["cycle"]
    cycle_id = conn.execute(
        "INSERT INTO agro_crop_cycle(parcel_id, cycle_kind, crop, season_year, "
        "sow_date, crop_status, source, confidence) VALUES(?,?,?,?,?,?,?,?)",
        (parcel_id, c["cycle_kind"], c["crop"], c["season_year"], c["sow_date"],
         c["crop_status"], c["source"], c["confidence"])).lastrowid
    invalid = []
    for e in rec["events"]:
        errs = _profiles.validate_event_attrs(e["event_type"], e["attrs"])
        if errs:
            invalid.append({"event_type": e["event_type"], "errors": errs})
        conn.execute(
            "INSERT INTO agro_event(parcel_id, cycle_id, season_year, event_type, "
            "attrs, source, confidence) VALUES(?,?,?,?,?,?,?)",
            (parcel_id, cycle_id, e["season_year"], e["event_type"],
             json.dumps(e["attrs"], ensure_ascii=False), e["source"], e["confidence"]))
    conn.commit()
    return {"is_vineyard": True,
            "written": {"parcel": 1, "cycle": 1, "events": len(rec["events"])},
            "invalid": invalid, "meta": parsed["meta"],
            "substances": len(parsed["substances"])}


# Обратная совместимость с прежней заглушкой (ADR-006 ранний контракт).
def parse_techcard(path: str | Path) -> dict[str, Any]:
    """Алиас: разобрать техкарту → {is_vineyard, meta, operations, substances}."""
    return parse_workbook(path)
