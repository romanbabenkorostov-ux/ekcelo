"""
egrn_parser/parsers/osv_parser.py — парсер оборотно-сальдовой ведомости (ОСВ).

Формат 1С, Excel, счёт 01 (Основные средства):
  01.01 — объекты в собственности (БУ-строки)
  01.К  — арендованные объекты (НУ-строки, ФСБУ 25)

Результат:
  - Список принадлежностей (accessories) — если режим включён
  - Список оценок (valuations) — всегда
  - Метаданные организации
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

from egrn_parser.parsers._common import (
    CAD_NUMBER_RE,
    parse_date_any,
    parse_number,
    normalize_whitespace,
    is_absent,
)
from egrn_parser.dictionaries import OSV_ACCOUNT_RIGHTS

log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
#  Константы
# ─────────────────────────────────────────────────────────────────────────────

# Маркеры ОСВ (ТЗ раздел 4.1)
OSV_MARKERS = ("оборотно-сальдовая ведомость", "01.01", "01.К", "бу", "ну")

QUARTER_MAP = {
    "1": ("01-01", "03-31"),
    "2": ("04-01", "06-30"),
    "3": ("07-01", "09-30"),
    "4": ("10-01", "12-31"),
}
HALF_MAP = {
    "1": ("01-01", "06-30"),
    "2": ("07-01", "12-31"),
}

# Частичный кадастровый номер (только последний сегмент)
CAD_PARTIAL_RE = re.compile(r":(\d{3,5})\b")

# Инвентарный номер
INV_RE_1 = re.compile(r"[Ии]нв\.?\s*№?\s*(\S+)")
INV_RE_2 = re.compile(r"\((\d{5,})\)")


# ─────────────────────────────────────────────────────────────────────────────
#  Определение файла ОСВ
# ─────────────────────────────────────────────────────────────────────────────

def is_osv_xlsx(path: Path | str) -> bool:
    """Проверить, является ли XLSX файлом ОСВ."""
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return False
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        for row_idx in range(1, 5):
            row = [str(ws.cell(row_idx, c).value or "").lower() for c in range(1, 5)]
            row_text = " ".join(row)
            if any(m in row_text for m in ("оборотно-сальдовая ведомость", "01.01", "01.к")):
                wb.close()
                return True
        wb.close()
        return False
    except Exception:
        return False


# ─────────────────────────────────────────────────────────────────────────────
#  Парсинг периода
# ─────────────────────────────────────────────────────────────────────────────

def parse_osv_period(title: str) -> tuple[str, str]:
    """
    Извлечь период из заголовка ОСВ.
    Поддерживаемые форматы:
      «за 1 квартал 2026 г»
      «за 1 полугодие 2026 г»
      «за 9 месяцев 2026 г»
      «за 2026 г»
    """
    # Квартал
    m = re.search(r"за\s+(\d)\s+квартал\s+(\d{4})", title, re.IGNORECASE)
    if m:
        q, y = m.group(1), m.group(2)
        if q in QUARTER_MAP:
            return f"{y}-{QUARTER_MAP[q][0]}", f"{y}-{QUARTER_MAP[q][1]}"

    # Полугодие
    m = re.search(r"за\s+(\d)\s+полугодие\s+(\d{4})", title, re.IGNORECASE)
    if m:
        h, y = m.group(1), m.group(2)
        if h in HALF_MAP:
            return f"{y}-{HALF_MAP[h][0]}", f"{y}-{HALF_MAP[h][1]}"

    # 9 месяцев
    m = re.search(r"за\s+9\s+месяцев\s+(\d{4})", title, re.IGNORECASE)
    if m:
        return f"{m.group(1)}-01-01", f"{m.group(1)}-09-30"

    # Год
    m = re.search(r"за\s+(\d{4})\s+г", title, re.IGNORECASE)
    if m:
        return f"{m.group(1)}-01-01", f"{m.group(1)}-12-31"

    raise ValueError(f"Не удалось распознать период ОСВ: {title!r}")


def _annualize_factor(period_from: str, period_to: str) -> float:
    """Коэффициент для пересчёта оборотов за период в годовые."""
    try:
        dt1 = datetime.strptime(period_from, "%Y-%m-%d")
        dt2 = datetime.strptime(period_to, "%Y-%m-%d")
        days = (dt2 - dt1).days + 1
        return round(365.0 / days, 6) if days > 0 else 1.0
    except Exception:
        return 1.0


# ─────────────────────────────────────────────────────────────────────────────
#  Извлечение кад. номера и инв. номера из наименования
# ─────────────────────────────────────────────────────────────────────────────

def extract_cad_from_name(name: str) -> tuple[Optional[str], Optional[str]]:
    """
    Извлечь кадастровый номер из наименования объекта ОС.
    Возвращает (full_cad_number, cad_fragment).
    """
    m = CAD_NUMBER_RE.search(name)
    if m:
        return m.group(1), None
    m = CAD_PARTIAL_RE.search(name)
    if m:
        return None, ":" + m.group(1)
    return None, None


def extract_inventory_number(name: str) -> Optional[str]:
    """Извлечь инвентарный номер из наименования объекта ОС."""
    m = INV_RE_1.search(name)
    if m:
        return m.group(1)
    m = INV_RE_2.search(name)
    if m:
        return m.group(1)
    return None


def _looks_like_datetime(value: Any) -> bool:
    """Проверить, является ли значение датой (маркер субсчёта ОСВ)."""
    if value is None:
        return False
    if isinstance(value, datetime):
        return True
    s = str(value).strip()
    return bool(re.match(r"\d{4}-\d{2}-\d{2}", s) or re.match(r"\d{2}\.\d{2}\.\d{4}", s))


# ─────────────────────────────────────────────────────────────────────────────
#  Главная функция парсинга
# ─────────────────────────────────────────────────────────────────────────────

def parse_osv_xlsx(
    xlsx_path: Path | str,
    include_accessories: bool = True,
) -> dict:
    """
    Разобрать ОСВ (оборотно-сальдовую ведомость) из Excel-файла 1С.

    Возвращает dict:
      entity_name:   str — наименование организации
      entity_inn:    str | None
      period_from:   str (YYYY-MM-DD)
      period_to:     str (YYYY-MM-DD)
      accessories:   list[dict] — если include_accessories=True
      valuations:    list[dict] — всегда
      warnings:      list[str]
    """
    xlsx_path = Path(xlsx_path)
    log.debug("Парсинг ОСВ: %s", xlsx_path.name)

    result: dict[str, Any] = {
        "entity_name":  None,
        "entity_inn":   None,
        "period_from":  None,
        "period_to":    None,
        "accessories":  [],
        "valuations":   [],
        "warnings":     [],
        "source_file":  xlsx_path.name,
    }

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error("Не удалось открыть ОСВ %s: %s", xlsx_path.name, e)
        result["warnings"].append(f"Ошибка открытия файла: {e}")
        return result

    # ── Шаг 1: Метаданные (строки 1–2) ────────────────────────────────────────
    row1_val = str(ws.cell(1, 1).value or "").strip()
    row2_val = str(ws.cell(2, 1).value or "").strip()

    result["entity_name"] = row1_val if row1_val else None

    try:
        period_from, period_to = parse_osv_period(row2_val)
        result["period_from"] = period_from
        result["period_to"]   = period_to
    except ValueError as e:
        result["warnings"].append(str(e))
        log.warning("ОСВ %s: %s", xlsx_path.name, e)

    annualize = _annualize_factor(
        result["period_from"] or "2025-01-01",
        result["period_to"]   or "2025-12-31",
    )

    # ── Шаг 2: Поиск строки-заголовка для определения колонок ─────────────────
    col_debit_start  = 3   # C
    col_credit_start = 4   # D
    col_turn_debit   = 5   # E
    col_turn_credit  = 6   # F
    col_debit_end    = 7   # G
    col_credit_end   = 8   # H

    # ── Шаг 3: Обход строк данных ─────────────────────────────────────────────
    current_section = "01.01"
    current_item    = None

    found_account_01 = False

    for row_idx in range(7, ws.max_row + 1):
        col_a = str(ws.cell(row_idx, 1).value or "").strip()
        col_b = str(ws.cell(row_idx, 2).value or "").strip()

        # Пропустить строки ВР (временные разницы)
        if col_b == "ВР":
            continue

        # Конец данных
        if col_a.lower().startswith("итого"):
            break

        # Итоговая строка счёта 01
        if col_a == "01" and col_b == "БУ":
            found_account_01 = True
            continue

        # Маркер перехода к субсчёту 01.К
        if col_a == "01.К" and col_b in ("БУ", ""):
            current_section = "01.К"
            continue

        # Маркер дата (субсчёт-заголовок)
        if col_b == "БУ" and _looks_like_datetime(ws.cell(row_idx, 1).value):
            continue

        # НУ-строка итога
        if col_a == "" and col_b == "НУ" and current_section == "01.01":
            continue

        def _num(col: int) -> float:
            v = ws.cell(row_idx, col).value
            return float(v) if isinstance(v, (int, float)) else (parse_number(str(v)) or 0.0)

        # ── 01.01: БУ-строка → собственный объект ОС ─────────────────────────
        if col_b == "БУ" and col_a and current_section == "01.01":
            current_item = col_a
            full_cad, cad_frag = extract_cad_from_name(col_a)
            inv_num = extract_inventory_number(col_a)

            debit_start  = _num(col_debit_start)
            debit_end    = _num(col_debit_end)
            turn_credit  = _num(col_turn_credit)

            # Принадлежность
            if include_accessories:
                acc: dict[str, Any] = {
                    "item_name":         col_a,
                    "inventory_number":  inv_num,
                    "re_cad_number":     full_cad,
                    "re_object_class":   None,
                    "cad_number_fragment": cad_frag,
                    "entity_name":       result["entity_name"],
                    "entity_inn":        result["entity_inn"],
                    "period_from":       result["period_from"],
                    "period_to":         result["period_to"],
                    "account_code":      "01.01",
                    "right_category":    "right",
                    "right_type":        "Собственность",
                    "is_disposed":       1 if (debit_end < 1 and turn_credit > 0) else 0,
                    "disposed_date":     result["period_to"] if (debit_end < 1 and turn_credit > 0) else None,
                    "source_file":       xlsx_path.name,
                }
                if cad_frag:
                    result["warnings"].append(
                        f"Частичный кад. номер «{cad_frag}» в «{col_a}» — требуется ручная привязка"
                    )
                result["accessories"].append(acc)

            # Оценки
            if debit_start > 0:
                result["valuations"].append({
                    "cad_number":           full_cad,
                    "accessory_name":       col_a,
                    "inventory_number":     inv_num,
                    "valuation_type":       "initial",
                    "amount":               debit_start,
                    "doc_date":             result["period_from"],
                    "period_label":         f"{result['period_from']}–{result['period_to']}",
                    "source_file":          xlsx_path.name,
                    "source_type":          "osv",
                    "object_class":         "accessory",
                })
            if debit_end > 0:
                result["valuations"].append({
                    "cad_number":           full_cad,
                    "accessory_name":       col_a,
                    "inventory_number":     inv_num,
                    "valuation_type":       "residual",
                    "amount":               debit_end,
                    "doc_date":             result["period_to"],
                    "period_label":         f"{result['period_from']}–{result['period_to']}",
                    "source_file":          xlsx_path.name,
                    "source_type":          "osv",
                    "object_class":         "accessory",
                })
            if debit_end < 1 and turn_credit > 0:
                result["valuations"].append({
                    "cad_number":       full_cad,
                    "accessory_name":   col_a,
                    "inventory_number": inv_num,
                    "valuation_type":   "writeoff",
                    "amount":           turn_credit,
                    "doc_date":         result["period_to"],
                    "source_file":      xlsx_path.name,
                    "source_type":      "osv",
                    "object_class":     "accessory",
                })

        # ── 01.К: НУ-строка → арендованный объект ────────────────────────────
        elif col_b == "НУ" and col_a and current_section == "01.К":
            current_item = col_a
            full_cad, cad_frag = extract_cad_from_name(col_a)
            inv_num = extract_inventory_number(col_a)

            debit_start = _num(col_debit_start)
            debit_end   = _num(col_debit_end)
            turn_credit = _num(col_turn_credit)

            if include_accessories:
                acc = {
                    "item_name":          col_a,
                    "inventory_number":   inv_num,
                    "re_cad_number":      full_cad,
                    "re_object_class":    None,
                    "cad_number_fragment":cad_frag,
                    "entity_name":        result["entity_name"],
                    "entity_inn":         result["entity_inn"],
                    "period_from":        result["period_from"],
                    "period_to":          result["period_to"],
                    "account_code":       "01.К",
                    "right_category":     "encumbrance",
                    "right_type":         "Аренда",
                    "is_disposed":        0,
                    "source_file":        xlsx_path.name,
                }
                if cad_frag:
                    result["warnings"].append(
                        f"Частичный кад. номер «{cad_frag}» в «{col_a}» — требуется ручная привязка"
                    )
                result["accessories"].append(acc)

            # Оценки аренды
            if debit_start > 0:
                result["valuations"].append({
                    "cad_number":       full_cad,
                    "accessory_name":   col_a,
                    "inventory_number": inv_num,
                    "valuation_type":   "initial",
                    "amount":           debit_start,
                    "doc_date":         result["period_from"],
                    "source_file":      xlsx_path.name,
                    "source_type":      "osv",
                    "object_class":     "accessory",
                })
            if debit_end > 0:
                result["valuations"].append({
                    "cad_number":       full_cad,
                    "accessory_name":   col_a,
                    "inventory_number": inv_num,
                    "valuation_type":   "residual",
                    "amount":           debit_end,
                    "doc_date":         result["period_to"],
                    "source_file":      xlsx_path.name,
                    "source_type":      "osv",
                    "object_class":     "accessory",
                })
            if turn_credit > 0:
                lease_annual = turn_credit * annualize
                result["valuations"].append({
                    "cad_number":       full_cad,
                    "accessory_name":   col_a,
                    "inventory_number": inv_num,
                    "valuation_type":   "lease_annual",
                    "amount":           round(lease_annual, 2),
                    "doc_date":         result["period_to"],
                    "source_file":      xlsx_path.name,
                    "source_type":      "osv",
                    "notes":            f"Пересчёт из оборота за период (×{annualize})",
                    "object_class":     "accessory",
                })

    if not found_account_01 and not result["accessories"] and not result["valuations"]:
        result["warnings"].append("Счёт 01 не найден в файле ОСВ")

    wb.close()

    log.info(
        "✓ ОСВ %s: %d принадлежностей, %d оценок, %d предупреждений",
        xlsx_path.name,
        len(result["accessories"]),
        len(result["valuations"]),
        len(result["warnings"]),
    )
    return result

# ─────────────────────────────────────────────────────────────────────────────
#  Извлечение субъекта из имени файла ОСВ (Fix 40h)
# ─────────────────────────────────────────────────────────────────────────────

def extract_entity_from_osv_filename(filename: str) -> dict:
    """
    Извлечь информацию о субъекте права из имени файла ОСВ.
    
    Паттерны:
    - «ОСВ ООО Санаторий Сосновая Роща.xlsx» → ООО «Санаторий Сосновая Роща»
    - «ОСВ_ООО_ССРП 01.2025.xlsx» → ООО «ССРП»
    
    Субъект из имени файла = владелец имущества по счёту 01.01.
    Субъект из текста пояснений к 01.К = арендодатель (не владелец).
    """
    import re
    
    name = filename
    # Убрать расширение и дату
    name = re.sub(r'\.[a-z]{3,4}$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\d{2}\.\d{4}$|\d{4}$', '', name).strip()
    # Убрать префикс «ОСВ», «ОСВ_», «осв »
    name = re.sub(r'^(?:ОСВ[_\s]+|осв[_\s]+)', '', name, flags=re.IGNORECASE).strip()
    # Извлечь ИНН если есть
    inn = None
    m_inn = re.search(r'\b(\d{10}|\d{12})\b', name)
    if m_inn:
        inn = m_inn.group(1)
        name = name[:m_inn.start()].strip()
    # Нормализация org form
    from egrn_parser.parsers.pdf_parser import _normalize_org_name
    name = _normalize_org_name(name.replace('_', ' ')).strip()
    
    if not name or len(name) < 2:
        return {}
    
    return {
        "name":        name,
        "inn":         inn,
        "holder_type": "legal_entity",
        "source_file": filename,
    }
