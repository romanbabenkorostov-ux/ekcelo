"""
egrn_parser/parsers/xlsx_template_parser.py — чтение шаблона Assets.

Читает листы «Земельные участки» (A–U) и «Здания, сооружения» (A–AC)
без изменения структуры.

ВАЖНО: листы 1–2 шаблона заморожены. Парсер только ЧИТАЕТ данные.
       При экспорте xlsx_exporter дополняет копию шаблона справа.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

import openpyxl

log = logging.getLogger(__name__)

# Заголовки листа «Земельные участки» (A–U, 21 колонка)
LAND_SHEET_NAME = "Земельные участки"
LAND_COLUMNS = [
    "num",            # A — № п/п
    "inventory_num",  # B — Инв. №
    "name_accounting",# C — Наименование по бух.учету
    "name_egrn",      # D — Наименование по выписке из ЕГРН
    "address",        # E — Адрес
    "cad_number",     # F — Кадастровый номер
    "area",           # G — Площадь, кв.м
    "land_category",  # H — Категория земель
    "permitted_use",  # I — Вид разрешённого использования
    "group_common",   # J — Группы Компаний с одним Собственником
    "group",          # K — Группа Компаний
    "owner",          # L — Собственник
    "status",         # M — Статус
    "right_type",     # N — Вид права
    "extract_number", # O — № выписки
    "extract_date",   # P — Дата выписки
    "book_value",     # Q — Балансовая стоимость
    "book_date",      # R — Дата постановки на баланс
    "encumbrances",   # S — Информация об ограничениях/обременениях
    "cadastral_value",# T — Кадастровая стоимость
    "market_value",   # U — Рыночная стоимость
]

# Заголовки листа «Здания, сооружения» (A–AC, 29 колонок)
BUILDING_SHEET_NAME = "Здания, сооружения"
BUILDING_COLUMNS = [
    "num",             # A — № п/п
    "inventory_num",   # B — Инв. №
    "name_accounting", # C — Наименование по бух.учету
    "object_type",     # D — Вид объекта
    "name_egrn",       # E — Наименование по выписке
    "address",         # F — Адрес
    "cad_number",      # G — Кадастровый номер
    "area",            # H — Площадь, кв.м
    "purpose",         # I — Назначение
    "floors_total",    # J — Количество этажей (в т.ч. подземных)
    "underground_floors",# K — Количество подземных этажей
    "wall_material",   # L — Материал стен
    "year_built",      # M — Завершение строительства
    "year_used",       # N — Ввод в эксплуатацию
    "land_cad_number", # O — Кадастровый № ЗУ
    "group_common",    # P — Группы Компаний с одним Собственником
    "group",           # Q — Группа Компаний
    "owner",           # R — Собственник
    "status",          # S — Статус
    "right_type",      # T — Вид права
    "extract_number",  # U — № выписки
    "extract_date",    # V — Дата выписки
    "book_value_initial",# W — Первоначальная балансовая стоимость
    "book_date",       # X — Дата постановки на баланс
    "book_value_residual",# Y — Остаточная балансовая стоимость
    "book_date_residual", # Z — Дата формирования остаточной стоимости
    "encumbrances",    # AA — Информация об ограничениях/обременениях
    "cadastral_value", # AB — Кадастровая стоимость
    "market_value",    # AC — Рыночная стоимость
]


def is_assets_template(path: Path | str) -> bool:
    """Проверить, является ли XLSX файлом шаблона Assets."""
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return False
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return LAND_SHEET_NAME in sheet_names or BUILDING_SHEET_NAME in sheet_names
    except Exception:
        return False


def parse_xlsx_template(path: Path | str) -> dict:
    """
    Прочитать данные из шаблона Assets.
    Возвращает dict с ключами 'land_rows' и 'building_rows'.
    Не изменяет файл.
    """
    path = Path(path)
    result: dict[str, Any] = {
        "land_rows":     [],
        "building_rows": [],
        "source_file":   path.name,
    }

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        log.error("Не удалось открыть шаблон %s: %s", path.name, e)
        return result

    # Читать лист «Земельные участки»
    if LAND_SHEET_NAME in wb.sheetnames:
        ws = wb[LAND_SHEET_NAME]
        for row in ws.iter_rows(min_row=3, values_only=True):  # данные с 3-й строки
            if not any(row):
                continue
            row_dict = {col: val for col, val in zip(LAND_COLUMNS, row)}
            if row_dict.get("cad_number"):  # есть кадастровый номер
                result["land_rows"].append(row_dict)

    # Читать лист «Здания, сооружения»
    if BUILDING_SHEET_NAME in wb.sheetnames:
        ws = wb[BUILDING_SHEET_NAME]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            row_dict = {col: val for col, val in zip(BUILDING_COLUMNS, row)}
            if row_dict.get("cad_number"):
                result["building_rows"].append(row_dict)

    wb.close()
    log.info(
        "✓ Шаблон %s: %d ЗУ, %d зданий/сооружений",
        path.name, len(result["land_rows"]), len(result["building_rows"]),
    )
    return result
