"""
egrn_parser/exporters/xlsx_exporter.py — экспорт в XLSX (14 листов).

ТЗ раздел 11.2:
  Листы 1–2 (Земельные участки / Здания, сооружения) — расширяются справа
  от существующей структуры шаблона. Колонки A–U и A–AC НЕ ИЗМЕНЯЮТСЯ.
  Новые поля — начиная с V (для ЗУ) и AD (для зданий).

  Листы 3–14 — полностью генерируются из SQLite.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from egrn_parser.db.connection import get_connection

log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
#  Нормализация наименований правообладателей (Fix 15)
# ─────────────────────────────────────────────────────────────────────────────

_ORG_FORMS = [
    ("Общество с ограниченной ответственностью", "ООО"),
    ("Акционерное общество", "АО"),
    ("Публичное акционерное общество", "ПАО"),
    ("Закрытое акционерное общество", "ЗАО"),
    ("Открытое акционерное общество", "ОАО"),
    ("Государственное унитарное предприятие", "ГУП"),
    ("Муниципальное унитарное предприятие", "МУП"),
    ("Федеральное государственное унитарное предприятие", "ФГУП"),
    ("Публично-правовая компания", "ППК"),
    ("Государственное бюджетное учреждение", "ГБУ"),
    ("Некоммерческая организация", "НКО"),
    ("Индивидуальный предприниматель", "ИП"),
]


def _shorten_org_form(name: str) -> str:
    """Сократить «Общество с ограниченной ответственностью» → «ООО» и т.д."""
    if not name:
        return name
    result = name
    for long_form, short_form in _ORG_FORMS:
        result = result.replace(long_form, short_form)
        result = result.replace(long_form.upper(), short_form)
        result = result.replace(long_form.lower(), short_form)
    return result


def _normalize_holder_name(name: str) -> str:
    """Нормализовать имя для дедупликации: нижний регистр, без пробелов."""
    if not name:
        return ""
    short = _shorten_org_form(name)
    return short.lower().replace(" ", "").replace("_", "").replace('"', "").replace("«", "").replace("»", "")



# ─────────────────────────────────────────────────────────────────────────────
#  Стили
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="CCFFCC")   # светло-зелёный (Fix 23)
HEADER_FONT   = Font(name="Calibri", color="0D0D0D", bold=True, size=9)  # Calibri 9pt #0D0D0D
DATA_FONT     = Font(name="Calibri", size=9)
NEW_COL_FILL  = PatternFill("solid", fgColor="E8FFE8")   # зеленоватый — новые колонки
NEW_COL_FONT  = Font(name="Calibri", bold=True, size=9, color="006400")  # тёмно-зелёный
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")

def _fmt_date(val) -> str:
    """ISO → «20.12.2016 21:18:49» (Fix 29)."""
    if not val:
        return ""
    import re
    # ISO datetime: 2016-12-20T21:18:49 или 2016-12-20 21:18:49
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})[T ](\d{2}:\d{2}:\d{2})", str(val))
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)} {m.group(4)}"
    # Только дата: 2016-12-20
    m2 = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(val))
    if m2:
        return f"{m2.group(3)}.{m2.group(2)}.{m2.group(1)}"
    return str(val)



# ─────────────────────────────────────────────────────────────────────────────
#  Вспомогательные функции
# ─────────────────────────────────────────────────────────────────────────────

def _format_right_date(iso_str):
    """Преобразовать ISO дату/дату-время → «DD.MM.YYYY HH:MM:SS» (Fix 29)."""
    if not iso_str:
        return None
    import re as _r
    # YYYY-MM-DDTHH:MM:SS или YYYY-MM-DDTHH:MM:SS±HH:MM
    m = _r.match(r"(\d{4})-(\d{2})-(\d{2})(?:T(\d{2}):(\d{2}):(\d{2}))?", str(iso_str))
    if not m:
        return iso_str
    y, mo, d = m.group(1), m.group(2), m.group(3)
    if m.group(4):
        h, mi, s = m.group(4), m.group(5), m.group(6)
        return f"{d}.{mo}.{y} {h}:{mi}:{s}"
    return f"{d}.{mo}.{y}"


def _write_sheet_header(ws, title: str, columns: list[str], new_col_start: int = 9999) -> None:
    """Записать строку-заголовок листа."""
    ws.cell(1, 1, title).font = Font(bold=True, size=11)
    for i, col_name in enumerate(columns, 1):
        cell = ws.cell(2, i, col_name)
        if i >= new_col_start:
            cell.fill = NEW_COL_FILL
            cell.font = NEW_COL_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN  # center+center+wrap (Fix 23)
        cell.border = THIN_BORDER
    ws.freeze_panes = "A3"


def _write_row(ws, row_idx: int, values: list, new_col_start: int = 9999) -> None:
    """Записать строку данных."""
    for i, val in enumerate(values, 1):
        cell = ws.cell(row_idx, i, val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN
        if i >= new_col_start:
            cell.fill = NEW_COL_FILL


def _autowidth(ws, max_width: int = 50) -> None:
    """Подобрать ширину колонок по содержимому."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _json_to_text(s: Optional[str]) -> str:
    """Преобразовать JSON-строку в читаемый текст."""
    if not s:
        return ""
    try:
        obj = json.loads(s)
        if isinstance(obj, list):
            return "; ".join(str(x) for x in obj)
        return str(obj)
    except Exception:
        return str(s)


# ─────────────────────────────────────────────────────────────────────────────
#  Лист 1: Земельные участки
# ─────────────────────────────────────────────────────────────────────────────

LAND_BASE_COLS = [
    "№ п/п", "Инв. №", "Наименование по бух.учету", "Наименование по выписке из ЕГРН",
    "Адрес", "Кадастровый номер", "Площадь, кв.м", "Категория земель",
    "Вид разрешенного использования", "Группы Компаний с одним Собственником",
    "Группа Компаний", "Собственник", "Статус", "Вид права",
    "№ выписки из ЕГРН", "Дата выписки из ЕГРН", "Балансовая стоимость, руб.",
    "Дата постановки на баланс", "Информация об ограничениях (обременениях) права (при наличии)",
    "Кадастровая стоимость, руб", "Рыночная стоимость, руб.",
]  # A–U (21 колонка)

LAND_EXT_COLS = [
    "Кадастровый квартал",             # V
    "Этажность зданий на ЗУ",          # W
    "Код ВРИ",                          # X
    "Ограничения объекта (сводка)",    # Y
    "Кол-во ограничений объекта",      # Z
    "Ограничения объекта (JSON)",       # AA
    "Последняя выписка №",             # AB
    "Дата последней выписки",          # AC
    "Геометрия (WKT)",                 # AD
    "Широта центроида",                # AE
    "Долгота центроида",               # AF
    "Дата кадастровой стоимости",      # AG
    "Контент-хеш",                     # AH
    "Источник",                        # AI (Fix 32)
]


def _export_land_objects(ws, conn, template_ws=None) -> None:
    all_cols = LAND_BASE_COLS + LAND_EXT_COLS
    _write_sheet_header(ws, "Земельные участки", all_cols,
                        new_col_start=len(LAND_BASE_COLS) + 1)

    # Если есть данные из шаблона — копируем строки 3+ из шаблона, иначе пустые колонки A–U
    rows = conn.execute("""
        SELECT lo.*,
               (SELECT e.extract_date FROM extracts e WHERE e.cad_number = lo.cad_number ORDER BY e.extract_date DESC LIMIT 1) AS last_extract_date,
               (SELECT e.extract_number FROM extracts e WHERE e.cad_number = lo.cad_number ORDER BY e.extract_date DESC LIMIT 1) AS last_extract_number,
               (SELECT GROUP_CONCAT(rh.name, '; ') FROM rights r JOIN right_holders rh ON rh.right_id = r.right_id WHERE r.object_key_value = lo.cad_number AND r.right_category = 'right' AND r.is_active = 1 LIMIT 1) AS owner_name,
               (SELECT r.right_type FROM rights r WHERE r.object_key_value = lo.cad_number AND r.right_category = 'right' AND r.is_active = 1 LIMIT 1) AS right_type,
               (SELECT GROUP_CONCAT(r.right_type || ': ' || COALESCE(r.beneficiary_name, ''), '; ') FROM rights r WHERE r.object_key_value = lo.cad_number AND r.right_category IN ('encumbrance','restriction') AND r.is_active = 1) AS encumbrances_text,
               (SELECT v.amount FROM valuations v WHERE v.cad_number = lo.cad_number AND v.valuation_type = 'initial' ORDER BY v.doc_date DESC LIMIT 1) AS book_value,
               (SELECT geog.geom_wkt FROM object_geometries geog WHERE geog.cad_number = lo.cad_number AND geog.is_current = 1 LIMIT 1) AS geom_wkt,
               (SELECT a.item_name FROM accessories a
                WHERE a.re_cad_number = lo.cad_number AND a.account_code = '01.01' LIMIT 1
               ) AS name_accounting,
               lo.data_source AS source_file
        FROM land_objects lo
        WHERE lo.lifecycle_status != 'deregistered'
        ORDER BY lo.cad_number
    """).fetchall()

    for row_idx, row in enumerate(rows, 3):
        r = dict(row)
        # Ограничения объекта — сводка
        obj_rest = r.get("object_restrictions") or "[]"
        rest_list = json.loads(obj_rest) if obj_rest else []
        rest_summary = "; ".join(
            x.get("description", x.get("type", "")) for x in rest_list
        ) if rest_list else ""

        vals = [
            row_idx - 2,                          # A — №
            None,                                  # B — Инв.№
            r.get("name_accounting"),              # C — Наименование по бух. из ОСВ (Fix 16)
            _json_to_text(r.get("permitted_uses")),# D — Наименование по выписке
            r.get("address"),                      # E — Адрес
            r.get("cad_number"),                   # F — Кадастровый номер
            r.get("area"),                         # G — Площадь
            r.get("land_category"),                # H — Категория земель
            _json_to_text(r.get("permitted_uses")),# I — ВРИ
            None, None,                            # J, K — Группы компаний
            r.get("owner_name"),                   # L — Собственник
            None,                                  # M — Статус
            r.get("right_type"),                   # N — Вид права
            r.get("last_extract_number"),          # O — № выписки
            r.get("last_extract_date"),            # P — Дата выписки
            r.get("book_value"),                   # Q — Балансовая стоимость
            None,                                  # R — Дата постановки на баланс
            r.get("encumbrances_text"),            # S — Ограничения/обременения ПРАВА
            r.get("cadastral_value"),              # T — Кадастровая стоимость
            None,                                  # U — Рыночная стоимость
            # Расширенные колонки V+:
            r.get("quarter_cad_number"),           # V — Квартал
            None,                                  # W — Этажность зданий на ЗУ
            None,                                  # X — Код ВРИ
            rest_summary,                          # Y — Ограничения объекта (сводка)
            len(rest_list),                        # Z — Кол-во ограничений
            obj_rest if rest_list else None,       # AA — JSON
            r.get("last_extract_number"),          # AB — Последняя выписка
            r.get("last_extract_date"),            # AC — Дата последней выписки
            r.get("geom_wkt"),                     # AD — WKT
            None,                                  # AE — Широта
            None,                                  # AF — Долгота
            r.get("cadastral_value_date"),         # AG — Дата кад. стоимости
            r.get("content_hash"),                 # AH — Хеш
            r.get("data_source"),                  # AI — Файл-источник (Fix 32)
        ]
        _write_row(ws, row_idx, vals, new_col_start=len(LAND_BASE_COLS) + 1)

    _autowidth(ws)
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(all_cols))}{len(rows) + 2}"


# ─────────────────────────────────────────────────────────────────────────────
#  Лист 2: Здания, сооружения
# ─────────────────────────────────────────────────────────────────────────────

BLDG_BASE_COLS = [
    "№ п/п", "Инв. №", "Наименование по бух.учету", "Вид объекта недвижимости",
    "Наименование по выписке из ЕГРН", "Адрес", "Кадастровый номер", "Площадь, кв.м",
    "Назначение", "Количество этажей (в т.ч. подземных)", "Количество подземных этажей",
    "Материал стен", "Завершение строительства", "Ввод в эксплуатацию",
    "Кадастровый № ЗУ, на котором расположен объект",
    "Группы Компаний с одним Собственником", "Группа Компаний", "Собственник",
    "Статус", "Вид права", "№ выписки из ЕГРН", "Дата выписки из ЕГРН",
    "Первоначальная балансовая стоимость, руб.", "Дата постановки на баланс",
    "Остаточная балансовая стоимость, руб.", "Дата формирования остаточной стоимости",
    "Информация об ограничениях (обременениях) права", "Кадастровая стоимость, руб",
    "Рыночная стоимость, руб.",
]  # A–AC (29 колонок)

BLDG_EXT_COLS = [
    "Этажей всего (по выписке)",      # AD
    "Этажей надземных",               # AE
    "Этажей подземных",               # AF
    "Этажность по осмотру",           # AG
    "Состояние по осмотру",           # AH
    "Кадастровый номер родителя",     # AI
    "Класс объекта родителя",         # AJ
    "Этажей надземных родителя",      # AK
    "Этажей подземных родителя",      # AL
    "Ограничения объекта (сводка)",   # AM
    "Кол-во ограничений объекта",     # AN
    "Ограничения объекта (JSON)",      # AO
    "Последняя выписка №",            # AP
    "Дата последней выписки",         # AQ
    "Геометрия (WKT)",                # AR
    "Широта центроида",               # AS
    "Долгота центроида",              # AT
    "Контент-хеш",                    # AU
    "Тип основной характеристики",    # AV  (для сооружений: "площадь застройки")
    "Значение осн. хар-ки",            # AW  main_value (Fix 34)
    "Ед. измерения осн. хар-ки",      # AX
    "Файл-источник",                   # AY  (Fix 32)
]


def _export_building_objects(ws, conn) -> None:
    all_cols = BLDG_BASE_COLS + BLDG_EXT_COLS
    _write_sheet_header(ws, "Здания, сооружения", all_cols,
                        new_col_start=len(BLDG_BASE_COLS) + 1)

    rows = conn.execute("""
        SELECT bo.*,
               (SELECT e.extract_date FROM extracts e WHERE e.cad_number = bo.cad_number ORDER BY e.extract_date DESC LIMIT 1) AS last_extract_date,
               (SELECT e.extract_number FROM extracts e WHERE e.cad_number = bo.cad_number ORDER BY e.extract_date DESC LIMIT 1) AS last_extract_number,
               (SELECT GROUP_CONCAT(rh.name, '; ') FROM rights r JOIN right_holders rh ON rh.right_id = r.right_id WHERE r.object_key_value = bo.cad_number AND r.right_category = 'right' AND r.is_active = 1 LIMIT 1) AS owner_name,
               (SELECT r.right_type FROM rights r WHERE r.object_key_value = bo.cad_number AND r.right_category = 'right' AND r.is_active = 1 LIMIT 1) AS right_type,
               (SELECT GROUP_CONCAT(r.right_type || COALESCE(': ' || r.beneficiary_name, ''), '; ') FROM rights r WHERE r.object_key_value = bo.cad_number AND r.right_category IN ('encumbrance','restriction') AND r.is_active = 1) AS encumbrances_text,
               (SELECT v.amount FROM valuations v WHERE v.cad_number = bo.cad_number AND v.valuation_type = 'initial' ORDER BY v.doc_date DESC LIMIT 1) AS book_value_initial,
               (SELECT v.amount FROM valuations v WHERE v.cad_number = bo.cad_number AND v.valuation_type = 'residual' ORDER BY v.doc_date DESC LIMIT 1) AS book_value_residual,
               (SELECT geog.geom_wkt FROM object_geometries geog WHERE geog.cad_number = bo.cad_number AND geog.is_current = 1 LIMIT 1) AS geom_wkt,
               (SELECT a.item_name FROM accessories a WHERE a.re_cad_number = bo.cad_number AND a.account_code = '01.01' LIMIT 1) AS name_accounting
        FROM building_objects bo
        WHERE bo.object_type IN ('building', 'structure', 'ons', 'complex')
          AND bo.lifecycle_status != 'deregistered'
        ORDER BY bo.cad_number
    """).fetchall()

    for row_idx, row in enumerate(rows, 3):
        r = dict(row)
        obj_rest = r.get("object_restrictions") or "[]"
        rest_list = json.loads(obj_rest) if obj_rest else []
        rest_summary = "; ".join(x.get("description", x.get("type", "")) for x in rest_list)

        land_cad = _json_to_text(r.get("land_cad_numbers"))

        vals = [
            row_idx - 2,
            None,                              # Инв.№
            r.get("name_accounting"),          # C — Наименование по бух.учету из ОСВ (Fix 16)
            r.get("object_type", ""),          # Вид ОН
            r.get("name"),                     # Наименование по выписке
            r.get("address"),
            r.get("cad_number"),
            r.get("area"),
            r.get("purpose"),
            r.get("floors_total"),
            r.get("underground_floors"),
            r.get("wall_material"),
            r.get("year_built"),
            r.get("year_used"),
            land_cad,                          # O — ЗУ-носители
            None, None,                        # P, Q — Группы компаний
            r.get("owner_name"),               # R — Собственник
            None,                              # S — Статус
            r.get("right_type"),               # T — Вид права
            r.get("last_extract_number"),      # U — № выписки
            r.get("last_extract_date"),        # V — Дата выписки
            r.get("book_value_initial"),       # W — Первоначальная балансовая
            None,                              # X — Дата постановки на баланс
            r.get("book_value_residual"),      # Y — Остаточная балансовая
            None,                              # Z — Дата форм. остаточной
            r.get("encumbrances_text"),        # AA — Ограничения ПРАВА
            r.get("cadastral_value"),          # AB — Кадастровая стоимость
            None,                              # AC — Рыночная стоимость
            # Расширенные колонки AD+:
            r.get("floors_total"),             # AD
            r.get("floors_above_ground"),      # AE
            r.get("underground_floors"),       # AF
            r.get("floors_inspection"),        # AG
            r.get("condition_inspection"),     # AH
            r.get("parent_cad_number"),        # AI
            r.get("parent_object_class"),      # AJ
            r.get("parent_floors_above_ground"),# AK
            r.get("parent_underground_floors"),# AL
            rest_summary,                      # AM
            len(rest_list),                    # AN
            obj_rest if rest_list else None,   # AO
            r.get("last_extract_number"),      # AP
            r.get("last_extract_date"),        # AQ
            r.get("geom_wkt"),                 # AR
            None,                              # AS — Широта
            None,                              # AT — Долгота
            r.get("content_hash"),             # AU
            r.get("main_char_type"),           # AV  (Fix 10)
            r.get("main_value"),               # AW  (Fix 34)
            r.get("main_unit"),                # AX
            r.get("data_source"),              # AY  (Fix 32)
        ]
        _write_row(ws, row_idx, vals, new_col_start=len(BLDG_BASE_COLS) + 1)

    _autowidth(ws)
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(all_cols))}{len(rows) + 2}"


# ─────────────────────────────────────────────────────────────────────────────
#  Листы 3–14: остальные данные
# ─────────────────────────────────────────────────────────────────────────────

def _export_simple(ws, title: str, headers: list[str], rows: list[dict]) -> None:
    _write_sheet_header(ws, title, headers)
    for i, row in enumerate(rows, 3):
        vals = [row.get(h.lower().replace(" ", "_").replace(".", "_")
                        .replace(",", "").replace("(", "").replace(")", ""))
                for h in headers]
        _write_row(ws, i, vals)
    _autowidth(ws)


def _export_rooms(ws, conn) -> None:
    """Помещения и машино-места (Fix 35: Инв.№, Наименование по бух.учету)."""
    _write_sheet_header(ws, "Помещения и машино-места",
                        ["№", "Инв. №", "Наименование по бух.учету", "Кадастровый номер",
                         "Тип", "Наименование по выписке из ЕГРН", "Адрес",
                         "Площадь", "Этаж", "Родитель", "Кадастровая стоимость",
                         "Вид права", "Этажей надземных (родитель)", "Этажей подземных (родитель)",
                         "Источник"])  # Fix 32
    rows = conn.execute("""
        SELECT bo.*,
               (SELECT r.right_type FROM rights r WHERE r.object_key_value = bo.cad_number
                AND r.right_category='right' AND r.is_active=1 LIMIT 1) AS right_type,
               (SELECT a.inventory_number FROM accessories a WHERE a.re_cad_number = bo.cad_number LIMIT 1) AS inv_num,
               (SELECT a.item_name FROM accessories a WHERE a.re_cad_number = bo.cad_number
                AND a.account_code='01.01' LIMIT 1) AS name_accounting
        FROM building_objects bo WHERE object_type IN ('room','parking')
        ORDER BY parent_cad_number, floor, cad_number
    """).fetchall()
    for i, row in enumerate(rows, 3):
        r = dict(row)
        _write_row(ws, i, [i-2, r.get("inv_num"), r.get("name_accounting"),
                            r.get("cad_number"), r.get("object_type"),
                            r.get("name"), r.get("address"), r.get("area"), r.get("floor"),
                            r.get("parent_cad_number"), r.get("cadastral_value"),
                            r.get("right_type"), r.get("parent_floors_above_ground"),
                            r.get("parent_underground_floors"), r.get("data_source")])
    _autowidth(ws)


def _export_rights(ws, conn, category: str, title: str) -> None:
    headers = ["№", "Кадастровый номер", "Категория", "Вид права", "Номер регистрации",
               "Дата регистрации", "Правообладатель", "ИНН правообладателя", "Доля",
               "Активно", "Бенефициар", "ИНН бенефициара", "Действует с", "Действует до",
               "Основание", "Источник"]  # Fix 29: +Источник
    _write_sheet_header(ws, title, headers)
    rows = conn.execute(
        """
        SELECT r.*,
               GROUP_CONCAT(rh.name, '; ') AS holder_name,
               GROUP_CONCAT(rh.inn, '; ')  AS holder_inn
        FROM rights r
        LEFT JOIN right_holders rh ON rh.right_id = r.right_id
        WHERE r.right_category = ?
        GROUP BY r.right_id
        ORDER BY r.object_key_value, r.right_date
        """,
        (category,),
    ).fetchall()
    for i, row in enumerate(rows, 3):
        r = dict(row)
        num   = r.get("share_numerator")
        den   = r.get("share_denominator")
        if not num and r.get("right_type", "").lower() in ("собственность", "ownership"):
            share_str = "1/1"
        else:
            share_str = f"{num}/{den}" if num else None
        # Дата в формате «DD.MM.YYYY HH:MM:SS» (Fix 29)
        right_date = _format_right_date(r.get("right_date"))
        valid_until = _format_right_date(r.get("valid_until"))
        valid_from  = _format_right_date(r.get("valid_from"))
        _write_row(ws, i, [i-2, r.get("object_key_value"), r.get("right_category"),
                            r.get("right_type"), r.get("right_number"), right_date,
                            r.get("holder_name"), r.get("holder_inn"), share_str,
                            "Да" if r.get("is_active") else "Нет",
                            r.get("beneficiary_name"), r.get("beneficiary_inn"),
                            valid_from, valid_until, r.get("basis"),
                            r.get("source_filename") or r.get("source_extract_number")])
    _autowidth(ws)


def _export_restrictions_sheet(ws, conn) -> None:
    """Лист «Ограничения» — объединяет ограничения объекта и ограничения прав."""
    headers = ["Вид", "Кадастровый номер", "№ права", "Реестровый №", "Тип (расшифровка)",
               "Тип документа-основания", "Номер документа", "Дата документа",
               "Дата начала", "Дата окончания", "ИНН бенефициара", "Бенефициар",
               "Примечания", "Файл-источник"]  # Fix 32
    _write_sheet_header(ws, "Ограничения", headers)

    # Часть 1: ограничения объектов (из land_objects + building_objects JSON)
    row_idx = 3
    for table in ("land_objects", "building_objects"):
        rows_obj = conn.execute(
            f"SELECT cad_number, object_restrictions FROM {table} WHERE object_restrictions IS NOT NULL"
        ).fetchall()
        for row in rows_obj:
            obj_rest = row["object_restrictions"] or "[]"
            try:
                rest_list = json.loads(obj_rest)
            except Exception:
                continue
            for restr in rest_list:
                bd = restr.get("basis_doc") or {}
                _write_row(ws, row_idx, [
                    "object", row["cad_number"], None,
                    restr.get("registry_number") or restr.get("type"), restr.get("description"),
                    bd.get("type"), bd.get("number"), bd.get("date"),
                    restr.get("valid_from"), restr.get("valid_to"),
                    None, None, restr.get("notes"),
                    restr.get("source_extract"),  # Fix 32
                ])
                row_idx += 1

    # Часть 2: ограничения прав
    rows_r = conn.execute(
        """
        SELECT right_number, object_key_value, right_type,
               basis, valid_from, valid_until,
               beneficiary_inn, beneficiary_name
        FROM rights WHERE right_category = 'restriction'
        """
    ).fetchall()
    for row in rows_r:
        r = dict(row)
        _write_row(ws, row_idx, [
            "right", r.get("object_key_value"), r.get("right_number"),
            None, r.get("right_type"),
            None, None, None,
            r.get("valid_from"), r.get("valid_until"),
            r.get("beneficiary_inn"), r.get("beneficiary_name"),
            r.get("basis"), r.get("source_extract_number"),  # Fix 32
        ])
        row_idx += 1
    _autowidth(ws)


def _export_code_dictionary(ws, conn) -> None:
    headers = ["Категория", "Код", "Значение (RU)", "Краткое", "Описание", "Активно"]
    _write_sheet_header(ws, "Словарь кодов", headers)
    rows = conn.execute(
        "SELECT category, code, value_ru, value_short, description, is_active "
        "FROM code_dictionary ORDER BY category, code"
    ).fetchall()
    for i, row in enumerate(rows, 3):
        r = dict(row)
        _write_row(ws, i, [r.get("category"), r.get("code"), r.get("value_ru"),
                            r.get("value_short"), r.get("description"),
                            "Да" if r.get("is_active") else "Нет"])
    _autowidth(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  Главная функция
# ─────────────────────────────────────────────────────────────────────────────

def export_xlsx(
    db_path: Path | str,
    out_path: Path | str,
    template_path: Optional[Path | str] = None,
) -> Path:
    """
    Экспортировать данные из SQLite в XLSX (14 листов).

    Если template_path задан — копирует его и дополняет листы 1–2 справа.
    Иначе создаёт новый файл.
    """
    db_path  = Path(db_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    import shutil
    if template_path and Path(template_path).exists():
        shutil.copy2(template_path, out_path)
        wb = openpyxl.load_workbook(str(out_path))
    else:
        wb = openpyxl.Workbook()
        # Удалить дефолтный лист
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_names = {
        1:  "Земельные участки",
        2:  "Здания, сооружения",
        3:  "Помещения и машино-места",
        4:  "ОНС",
        5:  "Принадлежности и оборудование",
        6:  "Бизнес-единицы",
        7:  "Права",
        8:  "Обременения",
        9:  "Ограничения",
        10: "Правообладатели",
        11: "События объектов",
        12: "События прав",
        13: "Оценка стоимости",
        14: "Словарь кодов",
        15: "Контакты",
        16: "Фото",
    }

    def _get_or_create_sheet(name: str):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(name)

    with get_connection(db_path, readonly=True) as conn:
        # Лист 1
        ws1 = _get_or_create_sheet("Земельные участки")
        _export_land_objects(ws1, conn)

        # Лист 2
        ws2 = _get_or_create_sheet("Здания, сооружения")
        _export_building_objects(ws2, conn)

        # Лист 3 — Помещения и машино-места
        ws3 = _get_or_create_sheet("Помещения и машино-места")
        _export_rooms(ws3, conn)

        # Лист 4 — ОНС
        ws4 = _get_or_create_sheet("ОНС")
        _write_sheet_header(ws4, "ОНС",
                             ["№", "Инв. №", "Наименование по бух.учету",
                              "Кадастровый номер", "Наименование по выписке из ЕГРН",
                              "Адрес", "Площадь", "% готовности", "Проектируемое назначение",
                              "Кадастровая стоимость", "ЗУ-носитель", "Источник"])  # Fix 35
        rows_ons = conn.execute("""
            SELECT bo.*,
                   (SELECT a.inventory_number FROM accessories a WHERE a.re_cad_number = bo.cad_number LIMIT 1) AS inv_num,
                   (SELECT a.item_name FROM accessories a WHERE a.re_cad_number = bo.cad_number
                    AND a.account_code='01.01' LIMIT 1) AS name_accounting
            FROM building_objects bo WHERE object_type = 'ons' ORDER BY cad_number
        """).fetchall()
        for i, row in enumerate(rows_ons, 3):
            r = dict(row)
            _write_row(ws4, i, [i-2, r.get("inv_num"), r.get("name_accounting"),
                                  r.get("cad_number"), r.get("name"),
                                  r.get("address"), r.get("area"),
                                  r.get("construction_stage"), r.get("purpose"),
                                  r.get("cadastral_value"), r.get("land_cad_numbers"),
                                  r.get("data_source")])
        _autowidth(ws4)

        # Лист 5 — Принадлежности
        ws5 = _get_or_create_sheet("Принадлежности и оборудование")
        _write_sheet_header(ws5, "Принадлежности и оборудование",
                             ["№", "ИНН орг.", "Счёт", "Наименование", "Инв.№",
                              "Привязка к ОН", "Период с", "Период по",
                              "Балансовая стоимость", "Широта", "Долгота",
                              "Широта 2", "Долгота 2", "Полилиния WKT", "Примечание"])
        rows_acc = conn.execute(
            "SELECT * FROM accessories ORDER BY entity_inn, item_name"
        ).fetchall()
        for i, row in enumerate(rows_acc, 3):
            r = dict(row)
            _write_row(ws5, i, [i-2, r.get("entity_inn"), r.get("account_code"),
                                  r.get("item_name"), r.get("inventory_number"),
                                  r.get("re_cad_number"), r.get("period_from"), r.get("period_to"),
                                  None, r.get("lat"), r.get("lon"),
                                  r.get("lat2"), r.get("lon2"), r.get("geom_polyline"), None])
        _autowidth(ws5)

        # Лист 6 — Бизнес-единицы
        ws6 = _get_or_create_sheet("Бизнес-единицы")
        _write_sheet_header(ws6, "Бизнес-единицы",
                             ["№", "Наименование", "Тип", "Родительский объект",
                              "Статус", "Площадь, кв.м", "Этажи"])
        rows_bu = conn.execute(
            "SELECT * FROM business_units ORDER BY parent_cad_number, unit_name"
        ).fetchall()
        for i, row in enumerate(rows_bu, 3):
            r = dict(row)
            _write_row(ws6, i, [i-2, r.get("unit_name"), r.get("unit_type"),
                                  r.get("parent_cad_number"), r.get("status"),
                                  r.get("area_sqm"), r.get("floors_occupied")])
        _autowidth(ws6)

        # Лист 7 — Права
        ws7 = _get_or_create_sheet("Права")
        _export_rights(ws7, conn, "right", "Права")

        # Лист 8 — Обременения
        ws8 = _get_or_create_sheet("Обременения")
        _export_rights(ws8, conn, "encumbrance", "Обременения")

        # Лист 9 — Ограничения
        ws9 = _get_or_create_sheet("Ограничения")
        _export_restrictions_sheet(ws9, conn)

        # Лист 10 — Правообладатели (Fix 15)
        ws10 = _get_or_create_sheet("Правообладатели")
        _write_sheet_header(ws10, "Правообладатели",
                             ["№", "ИНН", "ОГРН", "КПП", "Тип",
                              "Наименование (полное)", "Наименование (сокращ.)",
                              "Группа компаний", "Кад. номер"])
        rows_h = conn.execute(
            """SELECT rh.*, e.name_full AS entity_name, e.kpp AS entity_kpp, cg.group_name,
                      r.object_key_value AS cad_number
               FROM right_holders rh
               LEFT JOIN entity_registry e ON e.inn = rh.inn
               LEFT JOIN rights r ON r.right_id = rh.right_id
               LEFT JOIN company_groups cg ON cg.group_id = e.group_id
               ORDER BY COALESCE(rh.inn,''), rh.name"""
        ).fetchall()
        seen = set()
        row_idx = 3
        for row in rows_h:
            r = dict(row)
            inn = r.get("inn") or ""
            name_raw = r.get("name") or r.get("entity_name") or ""
            name_short = _shorten_org_form(name_raw)
            key = (inn, _normalize_holder_name(name_raw)) if inn else (name_raw.lower().strip(),)
            if key in seen:
                continue
            seen.add(key)
            _write_row(ws10, row_idx, [row_idx-2, inn, r.get("ogrn"), r.get("entity_kpp"),
                                        r.get("holder_type"), name_raw, name_short,
                                        r.get("group_name"), r.get("cad_number")])
            row_idx += 1
        _autowidth(ws10)

        # Лист 11 — События объектов
        ws11 = _get_or_create_sheet("События объектов")
        _write_sheet_header(ws11, "События объектов",
                             ["№", "Класс", "Кадастровый номер", "Seq", "Тип события",
                              "Дата", "Изменённые поля"])
        rows_oe = conn.execute(
            "SELECT * FROM object_events ORDER BY cad_number, event_seq"
        ).fetchall()
        for i, row in enumerate(rows_oe, 3):
            r = dict(row)
            _write_row(ws11, i, [i-2, r.get("object_class"), r.get("cad_number"),
                                   r.get("event_seq"), r.get("event_type"), r.get("event_date"),
                                   r.get("changed_fields")])
        _autowidth(ws11)

        # Лист 12 — События прав
        ws12 = _get_or_create_sheet("События прав")
        _write_sheet_header(ws12, "События прав",
                             ["№", "№ права", "Seq", "Тип события", "Дата"])
        rows_re = conn.execute(
            "SELECT * FROM right_events ORDER BY right_number, event_seq"
        ).fetchall()
        for i, row in enumerate(rows_re, 3):
            r = dict(row)
            _write_row(ws12, i, [i-2, r.get("right_number"), r.get("event_seq"),
                                   r.get("event_type"), r.get("event_date")])
        _autowidth(ws12)

        # Лист 13 — Оценка стоимости
        ws13 = _get_or_create_sheet("Оценка стоимости")
        _write_sheet_header(ws13, "Оценка стоимости",
                             ["№", "Класс", "Кадастровый номер", "Наименование ОС",
                              "Тип оценки", "Сумма", "Валюта", "Дата", "Источник"])
        rows_v = conn.execute(
            "SELECT * FROM valuations ORDER BY cad_number, valuation_type"
        ).fetchall()
        for i, row in enumerate(rows_v, 3):
            r = dict(row)
            _write_row(ws13, i, [i-2, r.get("object_class"), r.get("cad_number"),
                                   r.get("accessory_name"), r.get("valuation_type"),
                                   r.get("amount"), r.get("currency", "RUB"),
                                   r.get("doc_date"), r.get("source_file")])
        _autowidth(ws13)

        # Лист 14 — Словарь кодов
        ws14 = _get_or_create_sheet("Словарь кодов")
        _export_code_dictionary(ws14, conn)

        # Лист 15 — Контакты (Fix 14)
        ws15 = _get_or_create_sheet("Контакты")
        _export_contacts(ws15, conn)

        # Лист 16 — Фото (Fix 24)
        ws16 = _get_or_create_sheet("Фото")
        _export_photos(ws16, conn)

    # Установить правильный порядок листов
    desired_order = list(sheet_names.values())
    sheets_in_wb = wb.sheetnames
    ordered = [s for s in desired_order if s in sheets_in_wb]
    remaining = [s for s in sheets_in_wb if s not in ordered]
    wb._sheets = [wb[s] for s in ordered + remaining]

    wb.save(str(out_path))
    log.info("✓ XLSX экспорт: %s (%d листов)", out_path.name, len(wb.sheetnames))
    return out_path


def _export_contacts(ws, conn) -> None:
    """Лист «Контакты» (Fix 14)."""
    headers = ["Роль", "Наименование заказчика", "ИНН заказчика", "КПП заказчика",
               "Контактное лицо заказчика", "Наименование исполнителя",
               "ИНН исполнителя", "КПП исполнителя", "Контактное лицо исполнителя",
               "Номер договора", "Дата договора", "Дата акта"]
    _write_sheet_header(ws, "Контакты", headers)
    try:
        rows = conn.execute(
            "SELECT role, customer_name, customer_inn, customer_kpp, customer_contact,"
            " executor_name, executor_inn, executor_kpp, executor_contact,"
            " contract_number, contract_date, act_date FROM contacts ORDER BY contact_id"
        ).fetchall()
    except Exception:
        rows = []
    # Если таблицы нет — показать предзаполненные строки
    if not rows:
        default_rows = [
            ("Субподряд идентификации", None, None, None, "Бабенко", None, None, None, None, None, None, None),
            ("Подряд идентификации",    None, None, None, None, None, None, None, None, None, None, None),
            ("Заказ идентификации",     None, None, None, None, None, None, None, None, None, None, None),
        ]
        for i, vals in enumerate(default_rows, 3):
            _write_row(ws, i, list(vals))
    else:
        for i, row in enumerate(rows, 3):
            _write_row(ws, i, list(row))
    _autowidth(ws)


def _export_photos(ws, conn) -> None:
    """Лист «Фото» (Fix 24)."""
    headers = ["№", "Имя файла", "Папка", "Широта", "Долгота",
               "Угол съёмки", "Высота", "Дата", "Кадастровый номер"]
    _write_sheet_header(ws, "Фото", headers)
    try:
        rows = conn.execute(
            "SELECT file_name, folder_path, latitude, longitude, bearing, "
            "altitude, taken_at, cad_number FROM photos ORDER BY taken_at"
        ).fetchall()
    except Exception:
        rows = []
    for i, row in enumerate(rows, 3):
        r = dict(row)
        _write_row(ws, i, [i-2, r.get("file_name"), r.get("folder_path"),
                            r.get("latitude"), r.get("longitude"), r.get("bearing"),
                            r.get("altitude"), r.get("taken_at"), r.get("cad_number")])
    _autowidth(ws)

