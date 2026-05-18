"""
egrn_parser/scripts/dossier.py — Формирование досье по кадастровому номеру.

Использование:
    python -m egrn_parser.scripts.dossier --db output/egrn.db --cad 90:25:020102:24
    python -m egrn_parser.scripts.dossier --db output/egrn.db --cad 90:25:020102:24 --format xlsx
    python -m egrn_parser.scripts.dossier --db output/egrn.db --cad 90:25:020102:24 --format json

Доступные SQL-представления (views):
  v_all_objects        — все объекты (ЗУ + ОКС) в одной таблице
  v_rights_full        — права с правообладателями
  v_lease_contracts    — договоры аренды (действующие и истёкшие)
  v_object_dossier     — сводное досье по объекту
  v_pledges_prohibitions — ипотеки, запреты, ограничения

Примеры SQL-запросов:
  -- Основная карточка объекта
  SELECT * FROM v_object_dossier WHERE cad_number = '90:25:020102:24';

  -- Все права на объект
  SELECT right_category, right_type, holder_name, holder_inn,
         share_str, valid_from, valid_until, source_file
  FROM v_rights_full
  WHERE cad_number = '90:25:020102:24'
  ORDER BY right_category, right_date;

  -- Действующие договоры аренды
  SELECT * FROM v_lease_contracts WHERE status = 'Действует';

  -- Объекты с запретами на регистрацию
  SELECT cad_number, object_name, right_type, basis
  FROM v_pledges_prohibitions
  WHERE right_type LIKE '%Запрещение%';
"""

from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path


def get_dossier(db_path: str, cad_number: str) -> dict:
    """Получить полное досье по кадастровому номеру."""
    import sqlite3
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    result = {
        "cad_number": cad_number,
        "sections":   {},
    }

    # 1. Основная карточка
    row = conn.execute(
        "SELECT * FROM v_object_dossier WHERE cad_number = ?", (cad_number,)
    ).fetchone()
    if row:
        result["sections"]["основное"] = dict(row)
    else:
        result["error"] = f"Объект {cad_number} не найден в БД"
        conn.close()
        return result

    # 2. Права
    rights = conn.execute(
        "SELECT right_category, right_type, holder_name, holder_inn, "
        "share_str, right_date, valid_from, valid_until, basis, source_file "
        "FROM v_rights_full WHERE cad_number = ? ORDER BY right_category, right_date",
        (cad_number,),
    ).fetchall()
    result["sections"]["права"] = [dict(r) for r in rights]

    # 3. Договоры аренды
    leases = conn.execute(
        "SELECT * FROM v_lease_contracts WHERE cad_number = ?", (cad_number,)
    ).fetchall()
    result["sections"]["аренда"] = [dict(r) for r in leases]

    # 4. Обременения/Запреты
    pledges = conn.execute(
        "SELECT * FROM v_pledges_prohibitions WHERE cad_number = ?", (cad_number,)
    ).fetchall()
    result["sections"]["обременения"] = [dict(r) for r in pledges]

    # 5. История (события)
    events = conn.execute(
        "SELECT event_type, event_date, changed_fields, notes "
        "FROM object_events WHERE cad_number = ? ORDER BY event_seq",
        (cad_number,),
    ).fetchall()
    result["sections"]["события"] = [dict(e) for e in events]

    # 6. Оценки стоимости
    vals = conn.execute(
        "SELECT valuation_type, amount, currency, doc_date, source_file "
        "FROM valuations WHERE cad_number = ? ORDER BY doc_date DESC",
        (cad_number,),
    ).fetchall()
    result["sections"]["оценка"] = [dict(v) for v in vals]

    conn.close()
    return result


def print_dossier_text(dossier: dict) -> None:
    """Красивый текстовый вывод досье."""
    cad = dossier["cad_number"]
    print(f"\n{'='*65}")
    print(f"  ДОСЬЕ ОБЪЕКТА: {cad}")
    print(f"{'='*65}")

    main = dossier["sections"].get("основное", {})
    if main:
        print(f"\n  Тип:              {main.get('object_type_ru','')}")
        print(f"  Наименование:     {main.get('name','')}")
        print(f"  Адрес:            {main.get('address','')}")
        print(f"  Площадь:          {main.get('area','')} м²")
        print(f"  Кад. стоимость:   {main.get('cadastral_value','')} руб.")
        print(f"  Статус:           {main.get('lifecycle_status_text','')}")
        print(f"  Собственники:     {main.get('owners','—')}")
        print(f"  Обременений:      {main.get('encumbrances_count',0)}")
        print(f"  Ограничений:      {main.get('restrictions_count',0)}")
        print(f"  ЗОУИТ/ОКН:        {main.get('object_restrictions_count',0)} зон")
        print(f"  Последняя выписка:{main.get('last_extract','—')}")

    rights = dossier["sections"].get("права", [])
    if rights:
        print(f"\n  Зарегистрированные права ({len(rights)}):")
        for r in rights:
            holder = f"{r.get('holder_name','?')} (ИНН {r.get('holder_inn','?')})"
            print(f"    [{r.get('right_category','')}] {r.get('right_type','')} "
                  f"| {holder} | доля {r.get('share_str','?')}")

    leases = dossier["sections"].get("аренда", [])
    if leases:
        print(f"\n  Договоры аренды ({len(leases)}):")
        for l in leases:
            print(f"    № {l.get('contract_number','?')} | {l.get('status','?')} "
                  f"| {l.get('lease_start','?')} → {l.get('lease_end','?')}")

    pledges = dossier["sections"].get("обременения", [])
    if pledges:
        print(f"\n  Обременения/Ограничения прав ({len(pledges)}):")
        for p in pledges:
            print(f"    {p.get('category_ru','')} {p.get('right_type','')} "
                  f"| {p.get('beneficiary_name') or '—'}")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Досье по кадастровому номеру")
    parser.add_argument("--db",  required=True, help="SQLite БД")
    parser.add_argument("--cad", required=True, help="Кадастровый номер")
    parser.add_argument("--format", default="text", choices=["text","json","xlsx"])
    args = parser.parse_args(argv)

    dossier = get_dossier(args.db, args.cad)

    if args.format == "json":
        print(json.dumps(dossier, ensure_ascii=False, indent=2, default=str))

    elif args.format == "xlsx":
        from pathlib import Path
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for section, data in dossier["sections"].items():
            ws = wb.create_sheet(section)
            if isinstance(data, list) and data:
                headers = list(data[0].keys())
                for j, h in enumerate(headers, 1):
                    ws.cell(1, j, h)
                for i, row in enumerate(data, 2):
                    for j, h in enumerate(headers, 1):
                        ws.cell(i, j, str(row.get(h, "")))
            elif isinstance(data, dict):
                for i, (k, v) in enumerate(data.items(), 1):
                    ws.cell(i, 1, k)
                    ws.cell(i, 2, str(v))
        out_path = Path(f"dossier_{args.cad.replace(':','_')}.xlsx")
        wb.save(str(out_path))
        print(f"Сохранено: {out_path}")

    else:
        print_dossier_text(dossier)


if __name__ == "__main__":
    main()
