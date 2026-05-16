#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pirushin_sosn_rocha_052_make_structure_v1.py

Создание единой иерархической JSON-структуры:
  Предприятие → Бизнес-единицы → Кадастровые объекты → Уровни/Помещения → Оборудование

Источники:
  • ОСВ (.xlsx) по счёту 01 из 1С — обязательный вход.
  • Существующий structure_*.json — опц. база для идемпотентного обогащения.
  • Каталог с JSON-выгрузками NSPD / ЕГРН (опц.) — для подтягивания кадастровых
    объектов, адресов, этажности.

Запуск: python pirushin_sosn_rocha_052_make_structure_v1.py
Python 3.13+, Windows 10.

Зависимости: openpyxl
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# ─── Цветной вывод (Windows-совместимо) ────────────────────────────────────
class C:
    G = "\033[92m"; R = "\033[91m"; Y = "\033[93m"; CY = "\033[96m"
    M = "\033[95m"; B = "\033[1m"; X = "\033[0m"


def cp(text: str = "", color: str = "") -> None:
    print(f"{color}{text}{C.X}" if color else text)


# ─── Константы и справочники ───────────────────────────────────────────────
ACCOUNT_RE = re.compile(r"^\d{2}\.(?:\d{2,3}|К)$")
CN_FULL_RE = re.compile(r"\b\d{2}:\d{2}:\d{2,8}:\d{1,8}(?:/\d+)?\b")
CN_TAIL_RE = re.compile(r"(?<!\d):(\d{2,8})\b")           # «:119», «:8654»
INV_RE     = re.compile(r"\b01[_\.]\d{2,5}(?:[_\.]\d{2,5})?\b")
DATE_RE    = re.compile(r"\b(\d{2}\.\d{2}\.\d{4})\b")
CONTRACT_NO_RE = re.compile(r"№\s*([\w\-/]+)", re.IGNORECASE)

OPF_MAP: dict[str, str] = {
    "ООО":  "Общество с ограниченной ответственностью",
    "АО":   "Акционерное общество",
    "ПАО":  "Публичное акционерное общество",
    "ОАО":  "Открытое акционерное общество",
    "ЗАО":  "Закрытое акционерное общество",
    "ИП":   "Индивидуальный предприниматель",
    "ГУП":  "Государственное унитарное предприятие",
    "МУП":  "Муниципальное унитарное предприятие",
    "АНО":  "Автономная некоммерческая организация",
    "ФГБУ": "Федеральное государственное бюджетное учреждение",
    "ФГУП": "Федеральное государственное унитарное предприятие",
    "НКО":  "Некоммерческая организация",
    "ТСЖ":  "Товарищество собственников жилья",
}

ACCOUNT_MAP: dict[str, dict] = {
    "01":    {"label": "Основные средства (свод)",                 "right": None},
    "01.01": {"label": "Основные средства на балансе организации", "right": "собственность"},
    "01.03": {"label": "Арендованные основные средства",           "right": "аренда"},
    "01.09": {"label": "Амортизация арендованных ОС",              "right": "аренда"},
    "01.К":  {"label": "Арендные платежи (накопление)",            "right": "аренда"},
}

# Категории бизнес-единиц по ключевым словам в названии ОС
BU_KEYWORDS: dict[str, list[str]] = {
    "Медицина / Лечение": [
        "аппарат", "криотерап", "ванна", "реабокс", "массаж", "ингал",
        "ультразвук", "электрофор", "магнит", "лазер", "стоматолог",
        "тренажёр медиц", "процедур",
    ],
    "Спа / Бассейн / Бани": [
        "сауна", "хаммам", "парная", "бассейн", "джакузи", "купель",
        "пресотерап", "softcooker",
    ],
    "Ресторан / Кухня / Бар": [
        "плита", "пароконвектомат", "холодильник", "морозильник", "пицц",
        "соус", "фритюр", "посудомоечн", "кофе", "напитков", "бар",
        "аквариум для устриц", "акваферма",
    ],
    "Спорт / Фитнес": [
        "эллипсоид", "беговая дорожка", "велотренажёр", "штанга",
        "гантел", "тренаж", "precor",
    ],
    "Жилые номера / Гостиница": [
        "торшер", "кровать", "шкаф", "тумба", "матрас", "комод",
        "прикроват", "телевизор", "минибар",
    ],
    "Офис / Управление": [
        "xerox", "принтер", "мфу", "сканер", "ноутбук", "монитор",
        "компьютер", "проектор",
    ],
    "Транспорт / Водная техника": [
        "автомобиль", "транспортное средство", "ford", "lada", "toyota",
        "гидроцикл", "лодка", "катер", "яхта", "квадроцикл",
    ],
    "Земля / Ландшафт": [
        "земельный участок", "ландшафт", "благоустр", "тротуар",
    ],
    "Религиозное / Декор": [
        "аналой", "икон", "свечн", "храм", "церков", "электрокамин",
    ],
    "Технические / Инженерия": [
        "генератор", "трансформатор", "котёл", "котел", "насос",
        "кондиционер", "вентил",
    ],
}

# Признаки «уличного» расположения оборудования
OUTDOOR_HINTS = (
    "генератор", "трансформатор", "бассейн", "гидроцикл", "лодка", "катер",
    "яхта", "транспортное средство", "автомобиль", "акваферма",
)

LEVEL_TYPES: list[dict] = [
    {"key": "floor",              "name": "Этаж"},
    {"key": "mansarda",           "name": "Мансарда"},
    {"key": "mezonin",            "name": "Мезонин"},
    {"key": "basement",           "name": "Подвал"},
    {"key": "ground_floor",       "name": "Цокольный этаж"},
    {"key": "added_floor",        "name": "Надстроенный этаж"},
    {"key": "technical_floor",    "name": "Технический этаж"},
    {"key": "attic",              "name": "Чердак"},
    {"key": "antresol",           "name": "Антресоль"},
    {"key": "svetelka",           "name": "Светёлка"},
    {"key": "semi_basement",      "name": "Полуподвал"},
    {"key": "antresol_basement",  "name": "Антресоль подвала"},
    {"key": "antresol_ground",    "name": "Антресоль цокольного этажа"},
    {"key": "attic_addon",        "name": "Чердачная надстройка"},
    {"key": "operable_roof",      "name": "Эксплуатируемая кровля"},
    {"key": "tech_underfloor",    "name": "Техническое подполье"},
    {"key": "interfloor",         "name": "Междуэтажное пространство"},
]

CATEGORY_TO_OBJ_TYPE = {
    "Земельные участки":                          "Земельный участок",
    "Здания":                                     "Здание",
    "Сооружения":                                 "Сооружение",
    "Помещения":                                  "Помещение",
    "Объекты незавершенного строительства":       "Объект незавершенного строительства",
}


# ─── Утилиты ───────────────────────────────────────────────────────────────
def sha1_short(s: str, n: int = 8) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:n]


def slugify(s: str) -> str:
    table = {
        "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"zh",
        "з":"z","и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o",
        "п":"p","р":"r","с":"s","т":"t","у":"u","ф":"f","х":"h","ц":"c",
        "ч":"ch","ш":"sh","щ":"sch","ъ":"","ы":"y","ь":"","э":"e","ю":"yu",
        "я":"ya",
    }
    out: list[str] = []
    for ch in s.lower():
        if ch.isalnum() and ord(ch) < 128:
            out.append(ch)
        elif ch in table:
            out.append(table[ch])
        elif ch.isspace() or ch in "-_":
            out.append("_")
    res = "".join(out)
    res = re.sub(r"_+", "_", res).strip("_")
    return res or "unknown"


def norm_ws(s: str | None) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace(" ", " ")).strip()


def to_num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def ask(prompt: str, default: str = "") -> str:
    suffix = f" [{default}]" if default else ""
    ans = input(f"{prompt}{suffix}: ").strip().strip('"').strip("'")
    return ans or default


def ask_yn(prompt: str, default: bool = True) -> bool:
    suf = " [Y/n]" if default else " [y/N]"
    a = input(f"{prompt}{suf}: ").strip().lower()
    if not a:
        return default
    return a in ("y", "yes", "д", "да")


# ─── Извлечение enterprise из шапки ────────────────────────────────────────
def parse_enterprise(header_row: str, period_row: str) -> dict:
    h = norm_ws(header_row)
    opf, opf_full = None, None
    # Сначала длинные ключи, затем короткие (чтобы «ФГБУ» не подменялось «ИП»)
    for key in sorted(OPF_MAP, key=len, reverse=True):
        full = OPF_MAP[key]
        if h.lower().startswith(full.lower()):
            opf, opf_full = key, full
            break
        if re.search(rf"\b{re.escape(key)}\b", h):
            opf, opf_full = key, full
            break

    name_full = h
    name_short = h
    if opf_full and h.lower().startswith(opf_full.lower()):
        name_short = h[len(opf_full):].strip(' "«»')
    elif opf:
        m = re.search(rf"\b{re.escape(opf)}\b\s*(.+)", h)
        if m:
            name_short = m.group(1).strip(' "«»')

    period_label = norm_ws(period_row)
    period = {"label": period_label, "from": None, "to": None}
    # Очень мягкое распознавание «1 квартал 2026», «за <дата>», «<дата>-<дата>»
    qm = re.search(r"(\d)\s*квартал[а-я]*\s*(\d{4})", period_label, re.I)
    if qm:
        q, year = int(qm.group(1)), int(qm.group(2))
        start_m = (q - 1) * 3 + 1
        end_m = start_m + 2
        last_day = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][end_m - 1]
        if end_m == 2 and (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)):
            last_day = 29
        period["from"] = f"{year}-{start_m:02d}-01"
        period["to"]   = f"{year}-{end_m:02d}-{last_day:02d}"

    slug = slugify(name_short or name_full)
    return {
        "id": f"ent_{slug}",
        "slug": slug,
        "opf": opf,
        "opf_full": opf_full,
        "name_short": name_short,
        "name_full": name_full,
        "inn": None,
        "ogrn": None,
        "kpp": None,
        "period": period,
    }


# ─── Парсинг ОСВ ───────────────────────────────────────────────────────────
SECTION_LABELS = {
    "БУ", "НУ", "ВР", "Арендованные ОС", "Основные средства", "Итого",
}


def parse_osv(xlsx_path: Path) -> tuple[dict, list[dict]]:
    """
    Возвращает (enterprise, equipment_records).
    equipment record:
      { account, name, inv_hint, cadastral_hints, contract_date,
        contract_no, amounts:{bu,nu,vr} }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header = ""
    period = ""
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            header = norm_ws(row[0])
        elif i == 2:
            period = norm_ws(row[0])
        rows.append(row)

    enterprise = parse_enterprise(header, period)

    # Группируем тройки строк по имени объекта
    current_account: str | None = None
    items: list[dict] = []
    current: dict | None = None

    for row in rows:
        if not row:
            continue
        c0 = norm_ws(row[0])
        c1 = norm_ws(row[1])     # БУ / НУ / ВР

        if not c0 and not c1:
            continue

        if c0 and ACCOUNT_RE.match(c0):
            current_account = c0
            current = None
            continue

        if c0 == "Итого":
            current = None
            continue

        # Начало нового объекта: в col 0 — имя, в col 1 — «БУ»
        if c0 and c1 == "БУ" and c0 not in SECTION_LABELS:
            if current_account is None:
                continue
            current = {
                "account": current_account,
                "name": c0,
                "amounts": {
                    "bu": _row_amounts(row),
                    "nu": {"open_dt":0,"open_kt":0,"turnover_dt":0,"turnover_kt":0,"close_dt":0,"close_kt":0},
                    "vr": {"open_dt":0,"open_kt":0,"turnover_dt":0,"turnover_kt":0,"close_dt":0,"close_kt":0},
                },
            }
            items.append(current)
            continue

        # Продолжение тройки
        if current is not None and not c0 and c1 in ("НУ", "ВР"):
            key = "nu" if c1 == "НУ" else "vr"
            current["amounts"][key] = _row_amounts(row)
            continue

    # Постобработка: извлечение подсказок из названий
    for it in items:
        name = it["name"]
        it["cadastral_hints"] = _extract_cadastral_hints(name)
        inv = INV_RE.search(name)
        it["inv_hint"] = inv.group(0) if inv else None
        d = DATE_RE.search(name)
        it["contract_date"] = d.group(1) if d else None
        n = CONTRACT_NO_RE.search(name)
        it["contract_no"] = n.group(1) if n else None
        it["right_type"] = (ACCOUNT_MAP.get(it["account"], {}) or {}).get("right")

    return enterprise, items


def _row_amounts(row) -> dict:
    """ОСВ: cols 0=name, 1=label(БУ/НУ/ВР), 2..7 = 6 числовых ячеек."""
    cells = list(row) + [None] * (8 - len(row))
    return {
        "open_dt":     to_num(cells[2]),
        "open_kt":     to_num(cells[3]),
        "turnover_dt": to_num(cells[4]),
        "turnover_kt": to_num(cells[5]),
        "close_dt":    to_num(cells[6]),
        "close_kt":    to_num(cells[7]),
    }


def _extract_cadastral_hints(name: str) -> list[str]:
    hints: list[str] = []
    seen = set()
    for m in CN_FULL_RE.findall(name):
        if m not in seen:
            seen.add(m); hints.append(m)
    for m in CN_TAIL_RE.findall(name):
        token = f":{m}"
        if token not in seen:
            seen.add(token); hints.append(token)
    return hints


# ─── Загрузка JSON-источников (NSPD / EGRN / прежний structure) ────────────
def load_json_safe(path: Path) -> dict | list | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        cp(f"  ⚠ Не удалось прочитать {path.name}: {e}", C.Y)
        return None


def load_nspd_objects(folder: Path) -> list[dict]:
    """
    Принимает каталог; читает все *.json вида session_export_*.json или
    одиночных <kn>.json от 01_parsing_nspd_gemini_v5.py.
    Возвращает список плоских словарей по каждому объекту:
      { cn, category, info: {...}, related: {...} }
    """
    if not folder.exists() or not folder.is_dir():
        return []
    out: list[dict] = []
    for jp in folder.rglob("*.json"):
        data = load_json_safe(jp)
        if not isinstance(data, dict):
            continue
        payload = data.get("data") if "data" in data and isinstance(data["data"], dict) else data
        for category, by_cn in (payload or {}).items():
            if not isinstance(by_cn, dict):
                continue
            for cn, info in by_cn.items():
                if not isinstance(info, dict):
                    continue
                related = info.get("Связанные объекты", {}) if isinstance(info, dict) else {}
                out.append({
                    "cn": cn,
                    "category": category,
                    "info": {k: v for k, v in info.items() if k != "Связанные объекты"},
                    "related": related if isinstance(related, dict) else {},
                })
    return out


# ─── Построение бизнес-единиц по эвристике ─────────────────────────────────
def classify_bu(name: str) -> str:
    nl = name.lower()
    for bu, kws in BU_KEYWORDS.items():
        for kw in kws:
            if kw in nl:
                return bu
    return "Прочее / Не классифицировано"


def is_outdoor(name: str) -> bool:
    nl = name.lower()
    return any(h in nl for h in OUTDOOR_HINTS)


# ─── Сборка cadastre_objects из подсказок + NSPD ───────────────────────────
def build_cadastre_objects(
    equipment_records: list[dict],
    nspd_items: list[dict],
) -> tuple[list[dict], dict[str, str]]:
    """
    cadastre_objects + hint→cad_id map.
    """
    # Сначала из NSPD — у нас уже полные КН и тип объекта
    cad_by_cn: dict[str, dict] = {}
    for ni in nspd_items:
        cn = ni["cn"]
        if cn in cad_by_cn:
            continue
        obj_type = CATEGORY_TO_OBJ_TYPE.get(ni["category"], "Объект")
        info = ni["info"]
        cad_by_cn[cn] = {
            "id": f"cad_{sha1_short(cn)}",
            "cadastral_number": cn,
            "object_type": info.get("Вид объекта недвижимости") or obj_type,
            "right_type": None,
            "address": info.get("Адрес") or info.get("Местоположение"),
            "area": info.get("Площадь, кв. м") or info.get("Площадь"),
            "parent_cadastral": None,
            "levels": [],
            "enrichment_status": "partial",
            "sources": ["nspd"],
            "_raw_info": info,
        }

    # Затем добираем из подсказок в ОСВ (если в NSPD не нашлись)
    hint_to_id: dict[str, str] = {}
    for rec in equipment_records:
        for hint in rec["cadastral_hints"]:
            stub_key: str | None = None
            if CN_FULL_RE.fullmatch(hint):
                stub_key = hint
            else:
                tail = hint.lstrip(":")
                for known in cad_by_cn:
                    if known.endswith(":" + tail):
                        stub_key = known
                        break
                if not stub_key:
                    stub_key = hint  # сохраним «хвост» как ключ-заглушку
            if stub_key in cad_by_cn:
                hint_to_id[hint] = cad_by_cn[stub_key]["id"]
                continue
            cad_by_cn[stub_key] = {
                "id": f"cad_{sha1_short(stub_key)}",
                "cadastral_number": stub_key,
                "object_type": "Земельный участок" if "земельн" in rec["name"].lower() else "Объект",
                "right_type": (ACCOUNT_MAP.get(rec["account"], {}) or {}).get("right"),
                "address": None,
                "area": None,
                "parent_cadastral": None,
                "levels": [],
                "enrichment_status": "stub",
                "sources": ["osv_hint"],
            }
            hint_to_id[hint] = cad_by_cn[stub_key]["id"]

    return list(cad_by_cn.values()), hint_to_id


# ─── Авто-уровни здания из этажности NSPD/ЕГРН ─────────────────────────────
def compute_levels_for_cadastre(cad: dict) -> list[dict]:
    info = cad.get("_raw_info", {}) or {}
    obj_type = (cad.get("object_type") or "").lower()
    if "земельный" in obj_type:
        return []
    above = info.get("Количество этажей") or info.get("Этажность")
    below = info.get("Количество подземных этажей")

    def to_int(v):
        try: return int(re.search(r"\d+", str(v)).group(0))
        except Exception: return 0

    above_n = to_int(above) if above else 0
    below_n = to_int(below) if below else 0
    levels = []
    n = 1
    # снизу вверх: сначала подземные (от глубины к 1), затем надземные
    for i in range(below_n, 0, -1):
        lid = f"lvl_{sha1_short(cad['cadastral_number'] + f'_b{i}')}"
        levels.append({
            "id": lid,
            "number": n,
            "type": "Подвал" if i > 1 else "Цокольный этаж",
            "label": f"Уровень {n}. {'Подвал' if i > 1 else 'Цокольный этаж'} {i}",
            "underground": True,
            "cadastral_source": cad["cadastral_number"],
            "confirmed": False,
        })
        n += 1
    for i in range(1, above_n + 1):
        lid = f"lvl_{sha1_short(cad['cadastral_number'] + f'_a{i}')}"
        levels.append({
            "id": lid,
            "number": n,
            "type": "Этаж",
            "label": f"Уровень {n}. Этаж {i}",
            "underground": False,
            "cadastral_source": cad["cadastral_number"],
            "confirmed": False,
        })
        n += 1
    return levels


# ─── Идемпотентный merge с предыдущим structure.json ───────────────────────
def merge_preserve_confirmed(old: dict, new: dict) -> dict:
    """
    Сохраняет confirmed=true поля и пользовательские правки старой структуры.
    Стратегия: по id; confirmed=true → старая запись побеждает целиком;
    confirmed=false → переписывается новой, но user-only поля (name BU и т.п.)
    сохраняются.
    """
    if not old:
        return new

    def index(lst):
        return {x["id"]: x for x in lst if isinstance(x, dict) and "id" in x}

    for section in ("business_units", "cadastre_objects", "premises",
                    "equipment", "lease_contracts"):
        old_idx = index(old.get(section, []))
        merged: list[dict] = []
        seen: set[str] = set()
        for nv in new.get(section, []):
            ov = old_idx.get(nv["id"])
            if ov and ov.get("confirmed") is True:
                merged.append(ov)
            elif ov:
                # сливаем: новые данные + сохранённые пользовательские правки
                combined = {**nv, **{k: v for k, v in ov.items()
                                     if k in ("name", "links", "confirmed")
                                     and v not in (None, "", [], {})}}
                merged.append(combined)
            else:
                merged.append(nv)
            seen.add(nv["id"])
        # сохраняем confirmed-объекты, которых не было в новой генерации
        for oid, ov in old_idx.items():
            if oid not in seen and ov.get("confirmed") is True:
                merged.append(ov)
        new[section] = merged
    return new


# ─── Сборка equipment + lease_contracts ────────────────────────────────────
def build_equipment(
    records: list[dict],
    eq_to_bu: dict[str, str],
    hint_to_cad_id: dict[str, str],
) -> tuple[list[dict], list[dict]]:
    equipment: list[dict] = []
    leases: list[dict] = []

    # Маппинг 01.К → 01.03 по нормализованному имени
    name_to_eq01_03: dict[str, str] = {}

    for rec in records:
        eq_id = f"eq_{sha1_short(rec['account'] + '|' + rec['name'])}"
        bu_id = eq_to_bu.get(eq_id)
        outdoor = is_outdoor(rec["name"])

        cad_id = None
        for hint in rec["cadastral_hints"]:
            if hint in hint_to_cad_id:
                cad_id = hint_to_cad_id[hint]
                break

        location_kind = "standalone"
        if cad_id and outdoor:
            location_kind = "land_point"
        elif cad_id:
            location_kind = "level"

        eq = {
            "id": eq_id,
            "name": rec["name"],
            "account": rec["account"],
            "right_type": rec["right_type"],
            "inv_number_hint": rec["inv_hint"],
            "cadastral_hints": rec["cadastral_hints"],
            "amounts": rec["amounts"],
            "links": {
                "business_unit_id": bu_id,
                "cadastre_id": cad_id,
                "premises_id": None,
                "level_id": None,
                "location_kind": location_kind,
            },
            "confirmed": False,
        }
        equipment.append(eq)

        if rec["account"] == "01.03":
            name_to_eq01_03[_lease_key(rec["name"])] = eq_id

    # Второй проход: 01.К → lease_contracts
    for rec in records:
        if rec["account"] != "01.К":
            continue
        key = _lease_key(rec["name"])
        linked_eq = name_to_eq01_03.get(key)
        lc_id = f"lc_{sha1_short(rec['name'])}"
        leases.append({
            "id": lc_id,
            "equipment_ids": [linked_eq] if linked_eq else [],
            "name_reference": rec["name"],
            "cadastral_hint": rec["cadastral_hints"][0] if rec["cadastral_hints"] else None,
            "contract_no": rec["contract_no"],
            "contract_date": rec["contract_date"],
            "rent_quarterly_nu": rec["amounts"]["nu"]["turnover_kt"],
            "linked_account": "01.К",
        })

    return equipment, leases


def _lease_key(name: str) -> str:
    """Ключ для сопоставления 01.К ↔ 01.03 (одинаковые имена объектов аренды)."""
    s = name.lower()
    s = re.sub(r"\bдог[а-я]*\s*\S*", "", s)
    s = re.sub(r"\s+от\s+\d{2}\.\d{2}\.\d{4}", "", s)
    s = re.sub(r"№\s*\S+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ─── Сборка business_units: приоритет — по адресам, затем интерактивно ────
def _address_short(addr: str | None) -> str | None:
    """Короткая форма адреса для имени BU."""
    if not addr:
        return None
    a = norm_ws(addr)
    # Убираем индексы, страну, область
    a = re.sub(r"\b\d{6}\b,?\s*", "", a)
    a = re.sub(r"^Россий[а-я]+ Федерация,?\s*", "", a, flags=re.I)
    a = re.sub(r"^Республика [^,]+,?\s*", "", a, flags=re.I)
    a = re.sub(r"\bобл[а-я]*\.?\s+[^,]+,\s*", "", a, flags=re.I)
    return a[:80] or None


def resolve_equipment_address(
    rec: dict,
    cadastre_objects: list[dict],
    hint_to_cad_id: dict[str, str],
) -> tuple[str | None, str | None]:
    """Возвращает (cad_id, address) для оборудования, если разрешим."""
    cad_by_id = {c["id"]: c for c in cadastre_objects}
    for hint in rec["cadastral_hints"]:
        cid = hint_to_cad_id.get(hint)
        if not cid:
            continue
        cad = cad_by_id.get(cid)
        if cad and cad.get("address"):
            return cid, cad["address"]
        if cad:
            return cid, None
    return None, None


def build_business_units(
    records: list[dict],
    cadastre_objects: list[dict],
    hint_to_cad_id: dict[str, str],
) -> tuple[list[dict], dict[str, str]]:
    """
    Двухфазная стратегия:
      Фаза А — группировка по адресам из cadastre_objects (через КН-подсказки).
      Фаза Б — оставшиеся (без адреса) предлагаются пользователю на создание BU
               с подсказкой по ключевым словам.
    Возвращает (business_units, eq_id → bu_id).
    """
    # ── Фаза А: address-based grouping ──────────────────────────────────
    addr_groups: dict[str, list[dict]] = {}
    eq_unassigned: list[dict] = []
    for rec in records:
        if rec["account"] == "01.К":
            continue  # арендные платежи в BU не входят
        _, addr = resolve_equipment_address(rec, cadastre_objects, hint_to_cad_id)
        if addr:
            addr_groups.setdefault(addr, []).append(rec)
        else:
            eq_unassigned.append(rec)

    bus: list[dict] = []
    eq_to_bu: dict[str, str] = {}  # eq_id (account|name sha1) → bu_id

    if addr_groups:
        cp("\n" + "═" * 70, C.B)
        cp("Фаза А. Бизнес-единицы по адресам кадастровых объектов", C.B)
        cp("═" * 70, C.B)
        cp("Для каждой группы — Enter оставить предложенное имя или ввести своё.")
        for addr, items in sorted(addr_groups.items(), key=lambda kv: -len(kv[1])):
            short = _address_short(addr) or addr
            cp(f"\n  Адрес: {short}", C.CY)
            cp(f"  Объектов ОС в этой группе: {len(items)}")
            cp("  Примеры: " + " · ".join(it["name"][:40] for it in items[:4]))
            suggested = short
            name = ask("  Имя бизнес-единицы", default=suggested)
            bu_id = f"bu_{slugify(name)}"
            bus.append({
                "id": bu_id,
                "name": name,
                "keywords": [],
                "address": addr,
                "cadastrals": sorted({
                    hint
                    for it in items
                    for hint in it["cadastral_hints"]
                    if CN_FULL_RE.fullmatch(hint)
                }),
                "parent_id": None,
                "source": "address",
                "confirmed": True,
            })
            for it in items:
                eq_id = f"eq_{sha1_short(it['account'] + '|' + it['name'])}"
                eq_to_bu[eq_id] = bu_id

    # ── Фаза Б: остаток — интерактивное создание BU по подсказкам ──────
    if eq_unassigned:
        cp("\n" + "═" * 70, C.B)
        cp("Фаза Б. Оборудование без привязки к адресу", C.B)
        cp("═" * 70, C.B)

        # Кластеризация по эвристике для подсказок пользователю
        by_cat: dict[str, list[dict]] = {}
        for rec in eq_unassigned:
            by_cat.setdefault(classify_bu(rec["name"]), []).append(rec)

        cp(f"Всего без адреса: {len(eq_unassigned)}. Подсказка по эвристике:")
        for cat, items in sorted(by_cat.items(), key=lambda kv: -len(kv[1])):
            cp(f"  • {cat:<35} {len(items):>4} шт.", C.CY)

        cp("\nДействия для каждой эвристической группы:")
        cp("  [Enter] — создать BU с предложенным именем")
        cp("  имя     — создать BU с заданным именем")
        cp("  -       — пропустить (оборудование останется без BU)")
        cp("  =bu_id  — присоединить к ранее созданной BU (введите её id)")

        existing_bu_ids = {b["id"]: b for b in bus}

        for cat, items in sorted(by_cat.items(), key=lambda kv: -len(kv[1])):
            cp(f"\n  Группа: {cat} ({len(items)} объектов)")
            cp("  Примеры: " + " · ".join(it["name"][:40] for it in items[:4]))
            cmd = ask("  Имя BU / '-' / '=bu_id'", default=cat)
            if cmd == "-":
                continue
            if cmd.startswith("="):
                target_id = cmd[1:].strip()
                if target_id not in existing_bu_ids:
                    cp(f"    ⚠ BU id «{target_id}» не найдена, создаю новую с именем {cat}", C.Y)
                    cmd = cat
                else:
                    for it in items:
                        eq_id = f"eq_{sha1_short(it['account'] + '|' + it['name'])}"
                        eq_to_bu[eq_id] = target_id
                    continue
            bu_id = f"bu_{slugify(cmd)}"
            if bu_id in existing_bu_ids:
                target_id = bu_id
            else:
                bus.append({
                    "id": bu_id,
                    "name": cmd,
                    "keywords": BU_KEYWORDS.get(cat, []),
                    "address": None,
                    "cadastrals": [],
                    "parent_id": None,
                    "source": "user_keyword",
                    "confirmed": True,
                })
                existing_bu_ids[bu_id] = bus[-1]
                target_id = bu_id
            for it in items:
                eq_id = f"eq_{sha1_short(it['account'] + '|' + it['name'])}"
                eq_to_bu[eq_id] = target_id

    return bus, eq_to_bu


# ─── MAIN ──────────────────────────────────────────────────────────────────
def main() -> None:
    cp("\n" + "═" * 70, C.B)
    cp("  ОСВ → ИЕРАРХИЯ JSON: предприятие · BU · кадастр · оборудование", C.B)
    cp("═" * 70 + "\n", C.B)

    osv_path_str = ask("Путь к .xlsx ОСВ счёта 01")
    if not osv_path_str:
        cp("Отмена.", C.R); return
    osv_path = Path(osv_path_str)
    if not osv_path.exists() or osv_path.suffix.lower() != ".xlsx":
        cp(f"Файл не найден или не .xlsx: {osv_path}", C.R); return

    prev_path_str = ask("Путь к существующему structure_*.json (опционально, Enter — пропустить)")
    prev_data: dict | None = None
    if prev_path_str:
        p = Path(prev_path_str)
        if p.exists():
            prev_data = load_json_safe(p)
            cp(f"  ✓ Загружено: {p.name}", C.G)
        else:
            cp(f"  ⚠ Файл не найден, продолжаю без merge: {p}", C.Y)

    nspd_dir_str = ask("Каталог с JSON-выгрузками NSPD/ЕГРН (опционально, Enter — пропустить)")
    nspd_items: list[dict] = []
    if nspd_dir_str:
        d = Path(nspd_dir_str)
        nspd_items = load_nspd_objects(d)
        cp(f"  ✓ Объектов из NSPD/ЕГРН: {len(nspd_items)}", C.G)

    cp("\n⏳ Парсинг ОСВ…", C.CY)
    enterprise, records = parse_osv(osv_path)
    cp(f"  предприятие: {enterprise['opf']} «{enterprise['name_short']}»", C.G)
    cp(f"  период:      {enterprise['period']['label']}", C.G)
    cp(f"  ОС-записей:  {len(records)}", C.G)

    by_acc: dict[str, int] = {}
    for r in records:
        by_acc[r["account"]] = by_acc.get(r["account"], 0) + 1
    for a, n in sorted(by_acc.items()):
        right = (ACCOUNT_MAP.get(a, {}) or {}).get("right") or "—"
        cp(f"    {a:<6} ({right:<14}): {n}", C.CY)

    cp("\n⏳ Сборка кадастровых объектов…", C.CY)
    cadastre_objects, hint_to_cad_id = build_cadastre_objects(records, nspd_items)
    for cad in cadastre_objects:
        if cad.get("_raw_info"):
            cad["levels"] = compute_levels_for_cadastre(cad)
            cad.pop("_raw_info", None)
    cp(f"  кадастровых объектов: {len(cadastre_objects)}", C.G)

    cp("\n⏳ Построение бизнес-единиц (приоритет — по адресам)…", C.CY)
    business_units, eq_to_bu = build_business_units(records, cadastre_objects, hint_to_cad_id)
    cp(f"  бизнес-единиц создано: {len(business_units)}", C.G)

    cp("\n⏳ Сборка оборудования и договоров аренды…", C.CY)
    equipment, lease_contracts = build_equipment(records, eq_to_bu, hint_to_cad_id)
    cp(f"  оборудование:  {len(equipment)}", C.G)
    cp(f"  lease записей: {len(lease_contracts)}", C.G)

    structure = {
        "meta": {
            "schema_version": "1.0",
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "source_files": [str(osv_path)],
            "run_id": sha1_short(datetime.now().isoformat() + str(osv_path), 12),
        },
        "enterprise": enterprise,
        "dictionaries": {
            "level_types":  LEVEL_TYPES,
            "opf_map":      OPF_MAP,
            "account_map":  ACCOUNT_MAP,
            "bu_keywords":  BU_KEYWORDS,
        },
        "business_units": business_units,
        "cadastre_objects": cadastre_objects,
        "premises": prev_data.get("premises", []) if prev_data else [],
        "equipment": equipment,
        "lease_contracts": lease_contracts,
    }

    if prev_data:
        cp("\n⏳ Идемпотентный merge с предыдущим JSON…", C.CY)
        structure = merge_preserve_confirmed(prev_data, structure)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = osv_path.parent / f"structure_{enterprise['slug']}_{ts}.json"
    out.write_text(
        json.dumps(structure, ensure_ascii=False, indent=4),
        encoding="utf-8",
    )

    cp("\n" + "═" * 70, C.B)
    cp("ГОТОВО", C.B + C.G)
    cp("═" * 70, C.B)
    cp(f"  Файл: {out}", C.CY)
    cp(f"  Записей: ОС={len(equipment)} · BU={len(business_units)} · "
       f"кадастр={len(cadastre_objects)} · аренда={len(lease_contracts)}", C.G)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        cp("\n\nПрервано пользователем (Ctrl+C).", C.R)
        sys.exit(1)
    except Exception as e:
        cp(f"\n✗ Непредвиденная ошибка: {e}", C.R)
        import traceback; traceback.print_exc()
        sys.exit(1)
